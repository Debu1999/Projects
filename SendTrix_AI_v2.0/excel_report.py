from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
 
from db import (
    get_connection,
    get_snapshot_by_upload
)
 
import os
import pandas as pd
 
 
# =========================================================
# GENERATE COMPARISON EXCEL
# =========================================================
 
def generate_comparison_excel(
    comparison_id,
    master_upload_id,
    target_upload_id
):
 
    wb = Workbook()
 
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
 
    # =====================================================
    # LOAD DATA
    # =====================================================
 
    old_data = get_snapshot_by_upload(master_upload_id)
    new_data = get_snapshot_by_upload(target_upload_id)
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    SELECT
        appser_number,
        field_name,
        old_value,
        new_value,
        change_type
    FROM comparison_changes
    WHERE comparison_id = ?
    """, (comparison_id,))
 
    changes = cursor.fetchall()
 
    conn.close()
 
    # =====================================================
    # SUMMARY SHEET
    # =====================================================
 
    summary_sheet = wb.create_sheet("Summary")
 
    summary_sheet["A1"] = "Comparison Summary"
    summary_sheet["A1"].font = Font(bold=True, size=14)
 
    added_count = len([
        x for x in changes
        if x[4] == "ADDED"
    ])
 
    modified_count = len([
        x for x in changes
        if x[4] == "MODIFIED"
    ])
 
    missing_count = len([
        x for x in changes
        if x[4] == "MISSING"
    ])
 
    summary_data = [
        ["Metric", "Count"],
        ["Added Applications", added_count],
        ["Modified Fields", modified_count],
        ["Missing Applications", missing_count],
        ["Total Master Apps", len(old_data)],
        ["Total New Apps", len(new_data)]
    ]
 
    for row in summary_data:
        summary_sheet.append(row)
 
    # =====================================================
    # CONSOLIDATED SHEET
    # =====================================================
 
    consolidated = wb.create_sheet("Consolidated_Data")
 
    # =====================================================
    # DYNAMIC HEADERS
    # =====================================================
 
    all_headers = set()
 
    for row in old_data.values():
        all_headers.update(row.keys())
 
    for row in new_data.values():
        all_headers.update(row.keys())
 
    all_headers = list(all_headers)
 
    # Ensure ASN stays first
    if "appser_number" in all_headers:
        all_headers.remove("appser_number")
 
    all_headers.insert(0, "appser_number")
 
    # Add status column
    all_headers.append("CHANGE_STATUS")
 
    consolidated.append(all_headers)
 
    # =====================================================
    # HEADER STYLE
    # =====================================================
 
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )
 
    for cell in consolidated[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
 
    # =====================================================
    # QUICK LOOKUP
    # =====================================================
 
    change_lookup = {}
 
    for change in changes:
 
        asn = change[0]
        change_type = change[4]
 
        if asn not in change_lookup:
            change_lookup[asn] = set()
 
        change_lookup[asn].add(change_type)
 
    # =====================================================
    # COLORS
    # =====================================================
 
    added_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )
 
    modified_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )
 
    missing_fill = PatternFill(
        start_color="F4CCCC",
        end_color="F4CCCC",
        fill_type="solid"
    )
 
    # =====================================================
    # CURRENT DATA
    # =====================================================
 
    for asn, row in new_data.items():
 
        statuses = change_lookup.get(asn, {"UNCHANGED"})
 
        if "ADDED" in statuses:
            final_status = "ADDED"
 
        elif "MODIFIED" in statuses:
            final_status = "MODIFIED"
 
        else:
            final_status = "UNCHANGED"
 
        excel_row = []
 
        for header in all_headers[:-1]:
            excel_row.append(row.get(header, ""))
 
        excel_row.append(final_status)
 
        consolidated.append(excel_row)
 
        current_row = consolidated.max_row
 
        # Highlight row
        if final_status == "ADDED":
 
            for cell in consolidated[current_row]:
                cell.fill = added_fill
 
        elif final_status == "MODIFIED":
 
            for cell in consolidated[current_row]:
                cell.fill = modified_fill
 
    # =====================================================
    # MISSING DATA
    # =====================================================
 
    for asn, row in old_data.items():
 
        if asn not in new_data:
 
            excel_row = []
 
            for header in all_headers[:-1]:
                excel_row.append(row.get(header, ""))
 
            excel_row.append("MISSING")
 
            consolidated.append(excel_row)
 
            current_row = consolidated.max_row
 
            for cell in consolidated[current_row]:
                cell.fill = missing_fill
 
    # =====================================================
    # AUTO WIDTH
    # =====================================================
 
    for column_cells in consolidated.columns:
 
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )
 
        consolidated.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 5, 50)
 
    # =====================================================
    # SAVE FILE
    # =====================================================
 
    os.makedirs("reports", exist_ok=True)
 
    file_path = f"reports/comparison_{comparison_id}.xlsx"
 
    wb.save(file_path)
 
    return file_path
 
 
# =========================================================
# FINAL APPROVED EXPORT
# =========================================================
 
def generate_final_approved_excel(comparison_id):
 
    conn = get_connection()
 
    query = """
    SELECT
        appser_number,
        field_name,
        old_value,
        new_value,
        change_type,
        approval_status
    FROM comparison_changes
    WHERE comparison_id = ?
    """
 
    df = pd.read_sql_query(
        query,
        conn,
        params=(comparison_id,)
    )
 
    conn.close()
 
    output_path = f"comparison_final_{comparison_id}.xlsx"
 
    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:
 
        df.to_excel(
            writer,
            sheet_name="Approved_Changes",
            index=False
        )
 
    return output_path

def create_upload_from_comparison(
    comparison_id,
    master_upload_id,
    target_upload_id
):
 
    from datetime import datetime, timezone
 
    from db import (
        get_connection,
        get_snapshot_by_upload,
        insert_snapshot,
        insert_raw_snapshot
    )
 
    # =========================================
    # LOAD DATA
    # =========================================
 
    old_data = get_snapshot_by_upload(
        master_upload_id
    )
 
    new_data = get_snapshot_by_upload(
        target_upload_id
    )
 
    conn = get_connection()
    cursor = conn.cursor()
    # =========================================
    # LOAD APPROVED CHANGES
    # =========================================
 
    cursor.execute("""
    SELECT
        appser_number,
        field_name,
        new_value,
        change_type,
        approval_status
    FROM comparison_changes
    WHERE comparison_id = ?
    """, (comparison_id,))
 
    comparison_rows = cursor.fetchall()
    # =========================================
    # BUILD APPROVED CHANGE LOOKUP
    # =========================================
 
    approved_changes = {}
 
    approved_added_apps = set()
    approved_missing_apps=set()
 
    for row in comparison_rows:
 
        asn = row[0]
        field_name = row[1]
        new_value = row[2]
        change_type = row[3]
        approval_status = row[4]
 
        # Skip non-approved rows
        if approval_status != "APPROVED":
            continue
 
        # -------------------------------------
        # APPROVED ADDED APPLICATION
        # -------------------------------------
 
        if change_type == "ADDED":
 
            approved_added_apps.add(asn)
 
            continue

        # -------------------------------------
        # APPROVED MISSING APPLICATION
        # -------------------------------------
        if change_type == "MISSING":
            approved_missing_apps.add(asn)
            continue
        # -------------------------------------
        # APPROVED MODIFIED FIELD
        # -------------------------------------
 
        if asn not in approved_changes:
            approved_changes[asn] = {}
 
        approved_changes[asn][field_name] = new_value
 
 
    # =========================================
    # CREATE NEW CONSOLIDATED UPLOAD
    # =========================================
 
    now = datetime.now(timezone.utc).isoformat()
 
    generated_name = (
        f"comparison_master_{comparison_id}.xlsx"
    )
 
    cursor.execute("""
    INSERT INTO uploads (
        file_name,
        stored_name,
        created_at,
        file_type,
        comparison_id,
        is_master
    )
    VALUES (?, ?, ?, ?, ?,?)
    """, (
        generated_name,
        generated_name,
        now,
        "CONSOLIDATED",
        comparison_id,
        0
    ))
 
    new_upload_id = cursor.lastrowid
    # =========================================
    # BUILD FINAL CONSOLIDATED DATA
    # =========================================
 
    final_data = {}
 
    # -----------------------------------------
    # START WITH OLD MASTER
    # -----------------------------------------
 
    for asn, row in old_data.items():
        if asn in approved_missing_apps:
            continue
 
        final_data[asn] = row.copy()
 
    # -----------------------------------------
    # APPLY APPROVED CHANGES ONLY
    # -----------------------------------------
 
    for asn, row in new_data.items():
 
        # =====================================
        # APPROVED NEW APPLICATION
        # =====================================
 
        if asn not in old_data:
 
            # only add if approved
            if asn in approved_added_apps:
 
                final_data[asn] = row.copy()
 
            continue
 
        # =====================================
        # EXISTING APPLICATION
        # =====================================
 
        if asn not in approved_changes:
            continue
 
        # apply approved field changes only
        for field, approved_value in (
            approved_changes[asn].items()
        ):
 
            final_data[asn][field] = approved_value
 

    # =========================================
    # INSERT INTO SNAPSHOT TABLES
    # =========================================
 
    for asn, row in final_data.items():
 
        # UI/TRACKING SNAPSHOT
        insert_snapshot(
            cursor,
            new_upload_id,
            row
        )
 
        # FULL RAW DATA
        insert_raw_snapshot(
            cursor,
            new_upload_id,
            row
        )
    conn.commit()
    conn.close()
 
    return new_upload_id

def generate_excel_from_upload(upload_id,comparison_id):
 
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
 
    from db import get_snapshot_by_upload
 
    import os
 
    data = get_snapshot_by_upload(upload_id)
    print("DOWNLOADING UPLOAD:",upload_id)
    print("TOTAL ROWS:",len(data))
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
    SELECT
    appser_number,
    COUNT(*) as change_count,
    MAX(change_type) as change_type,
    MAX(approval_status) as approval_status
    FROM comparison_changes
    WHERE comparison_id=?
    GROUP BY appser_number
    ORDER BY appser_number
    """,(comparison_id,))
    summary_rows = cursor.fetchall()
    conn.close()
 
 
    wb = Workbook()
 
    sheet = wb.active
 
    sheet.title = "Consolidated_Data"
    summary_sheet = wb.create_sheet("Comparison_Summary")
    summary_sheet.append([
    "ASN",
    "Changes",
    "Type",
    "Status"
    ])
 
    # =========================================
    # DYNAMIC HEADERS
    # =========================================
 
    # =========================================
    # PRESERVE ORIGINAL COLUMN ORDER
    # =========================================
 
    all_headers = []
 
    # Take first row column order
    first_row = next(iter(data.values()), {})
 
    for key in first_row.keys():
        all_headers.append(key)
 
    sheet.append(all_headers)
 
    # =========================================
    # HEADER STYLE
    # =========================================
 
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )
 
    for cell in sheet[1]:
 
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
 
        cell.fill = header_fill

    print("SUMMARY ROWS:")
    for r in summary_rows[:20]:
        print(r)
    for row in summary_rows:
        summary_sheet.append([
            row[0],  # ASN
            row[1],  # Change Count
            row[2],  # Type
            row[3]   # Status
        ])
 
    # =========================================
    # DATA ROWS
    # =========================================
 
    for asn, row in data.items():
 
        excel_row = []
 
        for header in all_headers:
 
            excel_row.append(
                row.get(header, "")
            )
 
        sheet.append(excel_row)
 
    # =========================================
    # AUTO WIDTH
    # =========================================
 
    for column_cells in sheet.columns:
 
        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )
 
        sheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 5, 50)
 
    # =========================================
    # SAVE
    # =========================================
 
    os.makedirs("reports", exist_ok=True)
 
    output_path = (
        f"reports/final_upload_{upload_id}.xlsx"
    )
 
    wb.save(output_path)
 
    return output_path
 

 