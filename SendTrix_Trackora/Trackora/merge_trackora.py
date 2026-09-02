"""
Trackora -> Pretty COTS Tracker Excel Merge Script
====================================================

Merges a Trackora "Master Analysis" export into the manually-maintained
"pretty" COTS Tracker Excel workbook.

Behavior:
  - Application Name, SecureTrack ID, SBG, Application Owner:
        diffed against current values; if changed, the change is
        APPLIED and a row is written to the "Activity Log" tab
        (old value -> new value), so nothing is silently overwritten
        without a trace.
  - Patch Release Frequency, Compliance Status, Date Marked:
        directly overwritten from the latest Trackora data, no logging.
  - Evidence Received / Evidence Date / Timestamp Valid (G/H/I):
        left completely untouched.
  - New ASNs present in Trackora but not yet in the tracker are
        appended as new rows and logged as "New/Unmatched ASN".
  - The COTS Summary box and SBG breakdown table are recalculated
        from the final state of the tracker after all changes.

Usage:
    pip install openpyxl
    python merge_trackora.py --pretty "COTS_Tracker.xlsx" --trackora "Trackora_Master_20260901.xlsx" --output "COTS_Tracker_Updated.xlsx"

    Optional: --source-label "Sept 2026 client dump" to customize how
    this merge run is labeled in the Activity Log (defaults to the
    Trackora filename + timestamp).
"""

import argparse
from datetime import datetime
from openpyxl import load_workbook

TRACKER_SHEET = "COTS Tracker"
ACTIVITY_LOG_SHEET = "Activity Log"
MASTER_ANALYSIS_SHEET = "Master Analysis"

# Column positions in the COTS Tracker sheet (1-indexed to match openpyxl)
COL_APP_NAME = 2              # B
COL_SECURETRACK = 3           # C
COL_SBG = 4                   # D
COL_OWNER = 5                 # E
COL_FREQUENCY = 6             # F
# Columns 7,8,9 (G,H,I - Evidence Received/Date/Timestamp Valid) skipped
COL_COMPLIANCE_STATUS = 10    # J
COL_DATE_MARKED = 11          # K

FIRST_DATA_ROW = 2


def bucket_frequency(frequency, unit):
    """Convert a numeric frequency + unit into Monthly/Quarterly/Biannually/Annually,
    or 'Needs Review' if it doesn't cleanly fit any standard cadence."""
    try:
        value = float(frequency)
    except (TypeError, ValueError):
        return "Needs Review"

    unit = (unit or "").strip().lower()
    if unit.startswith("minute"):
        days = value / 1440
    elif unit.startswith("hour"):
        days = value / 24
    elif unit.startswith("day"):
        days = value
    elif unit.startswith("week"):
        days = value * 7
    elif unit.startswith("month"):
        days = value * 30
    elif unit.startswith("year"):
        days = value * 365
    else:
        return "Needs Review"

    if 25 <= days <= 40:
        return "Monthly"
    if 75 <= days <= 110:
        return "Quarterly"
    if 150 <= days <= 200:
        return "Biannually"
    if 330 <= days <= 400:
        return "Annually"
    return "Needs Review"


def load_master_data(trackora_path):
    """Read the Trackora 'Master Analysis' sheet into a dict keyed by ASN."""
    wb = load_workbook(trackora_path, data_only=True)
    if MASTER_ANALYSIS_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{MASTER_ANALYSIS_SHEET}' not found in {trackora_path}. "
            f"Found sheets: {wb.sheetnames}"
        )
    ws = wb[MASTER_ANALYSIS_SHEET]

    headers = [cell.value for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    required = ["ASN", "Name", "SBG", "Owner", "Frequency", "Frequency Unit",
                "Start Date", "Internal Status"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(
            f"Expected column(s) {missing} not found in '{MASTER_ANALYSIS_SHEET}' sheet. "
            f"Found columns: {headers}"
        )

    master = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        asn = row[idx["ASN"]]
        if not asn:
            continue
        asn = str(asn).strip()
        master[asn] = {
            "name": row[idx["Name"]],
            "sbg": row[idx["SBG"]],
            "owner": row[idx["Owner"]],
            "frequency": row[idx["Frequency"]],
            "frequency_unit": row[idx["Frequency Unit"]],
            "start_date": row[idx["Start Date"]],
            "internal_status": row[idx["Internal Status"]],
        }
    return master


def find_label_cell(ws, label, max_row=200, max_col=30):
    """Search the sheet for a cell whose text matches `label` (case-insensitive, exact match)."""
    label_lower = label.strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == label_lower:
                return cell
    return None


def update_summary_tables(ws, tracker_rows):
    """Recompute the COTS Summary box and SBG breakdown table based on
    the final state of the tracker rows after merge."""
    total = len(tracker_rows)
    compliant = sum(1 for r in tracker_rows if (r["compliance_status"] or "").strip().lower() == "compliant")
    non_compliant = sum(1 for r in tracker_rows if (r["compliance_status"] or "").strip().lower() == "non-compliant")
    pending = sum(1 for r in tracker_rows if (r["compliance_status"] or "").strip().lower() == "pending")
    under_review = sum(1 for r in tracker_rows if (r["compliance_status"] or "").strip().lower() == "under review")
    compliance_fraction = (compliant / total) if total else 0

    def set_next_to_label(label, value):
        cell = find_label_cell(ws, label)
        if cell:
            ws.cell(row=cell.row, column=cell.column + 1, value=value)
        else:
            print(f"  [warn] Could not find summary label '{label}' in sheet — skipped.")

    set_next_to_label("Total Applications", total)
    set_next_to_label("Compliant", compliant)
    set_next_to_label("Non-Compliant", non_compliant)
    set_next_to_label("Pending", pending)
    set_next_to_label("Under Review", under_review)
    set_next_to_label("Compliance %", compliance_fraction)  # raw fraction; relies on
                                                              # the cell's existing % format

    # --- SBG breakdown table ---
    sbg_header_cell = find_label_cell(ws, "SBG")
    if not sbg_header_cell:
        print("  [warn] Could not find 'SBG' breakdown table header — skipped.")
        return

    header_row = sbg_header_cell.row
    header_col = sbg_header_cell.column
    col_map = {}
    for c in range(header_col, header_col + 6):
        val = ws.cell(row=header_row, column=c).value
        if val:
            col_map[str(val).strip().lower()] = c

    sbg_totals = {}
    for r in tracker_rows:
        sbg = (r["sbg"] or "").strip().upper()
        if not sbg:
            continue
        bucket = sbg_totals.setdefault(sbg, {"total": 0, "compliant": 0, "non_compliant": 0})
        bucket["total"] += 1
        status = (r["compliance_status"] or "").strip().lower()
        if status == "compliant":
            bucket["compliant"] += 1
        elif status == "non-compliant":
            bucket["non_compliant"] += 1

    row_ptr = header_row + 1
    while True:
        sbg_label_cell = ws.cell(row=row_ptr, column=header_col)
        if not sbg_label_cell.value:
            break
        sbg_code = str(sbg_label_cell.value).strip().upper()
        data = sbg_totals.get(sbg_code, {"total": 0, "compliant": 0, "non_compliant": 0})
        pct = (data["compliant"] / data["total"]) if data["total"] else 0

        if "total" in col_map:
            ws.cell(row=row_ptr, column=col_map["total"], value=data["total"])
        if "compliant" in col_map:
            ws.cell(row=row_ptr, column=col_map["compliant"], value=data["compliant"])
        if "non-compliant" in col_map:
            ws.cell(row=row_ptr, column=col_map["non-compliant"], value=data["non_compliant"])
        if "compliance" in col_map:
            ws.cell(row=row_ptr, column=col_map["compliance"], value=pct)

        row_ptr += 1


def get_or_create_activity_log(wb):
    if ACTIVITY_LOG_SHEET in wb.sheetnames:
        return wb[ACTIVITY_LOG_SHEET]
    ws = wb.create_sheet(ACTIVITY_LOG_SHEET)
    ws.append(["Date", "ASN", "Application Name", "Changes Summary", "Source Upload"])
    return ws


def main():
    parser = argparse.ArgumentParser(
        description="Merge a Trackora Master Analysis export into the pretty COTS Tracker Excel."
    )
    parser.add_argument("--pretty", required=True, help="Path to the pretty Excel tracker file")
    parser.add_argument("--trackora", required=True, help="Path to the Trackora Master Analysis download")
    parser.add_argument("--output", required=True, help="Path to save the updated file")
    parser.add_argument("--source-label", default=None,
                         help="Label for this merge run in the Activity Log "
                              "(default: trackora filename + timestamp)")
    args = parser.parse_args()

    source_label = args.source_label or (
        f"{args.trackora.split('/')[-1].split(chr(92))[-1]} @ {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    print(f"Loading Trackora data from: {args.trackora}")
    master = load_master_data(args.trackora)
    print(f"  Loaded {len(master)} applications from Trackora.")

    print(f"Loading pretty tracker from: {args.pretty}")
    wb = load_workbook(args.pretty)
    if TRACKER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TRACKER_SHEET}' not found. Found sheets: {wb.sheetnames}")
    ws = wb[TRACKER_SHEET]

    log_ws = get_or_create_activity_log(wb)
    log_row_ptr = log_ws.max_row + 1

    matched_asns = set()
    tracker_rows_for_summary = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    changes_logged = 0

    def diff_and_apply(row_ptr, col, current_value, new_value, field_label, changes_list):
        """Applies the change if different, and appends a description to changes_list
        instead of writing to the log immediately - caller writes one combined
        log row per ASN after all fields are checked."""
        new_value_str = "" if new_value is None else str(new_value).strip()
        current_value_str = "" if current_value is None else str(current_value).strip()
        if new_value_str and new_value_str != current_value_str:
            changes_list.append(f"{field_label}: {current_value_str} \u2192 {new_value_str}")
            ws.cell(row=row_ptr, column=col, value=new_value_str)

    def write_log_row(asn, display_name, summary_text):
        nonlocal log_row_ptr, changes_logged
        log_ws.cell(row=log_row_ptr, column=1, value=now_str)
        log_ws.cell(row=log_row_ptr, column=2, value=asn)
        log_ws.cell(row=log_row_ptr, column=3, value=display_name)
        log_ws.cell(row=log_row_ptr, column=4, value=summary_text)
        log_ws.cell(row=log_row_ptr, column=5, value=source_label)
        log_row_ptr += 1
        changes_logged += 1

    print("Processing existing tracker rows...")
    row_ptr = FIRST_DATA_ROW
    while True:
        asn_cell = ws.cell(row=row_ptr, column=COL_SECURETRACK)
        if not asn_cell.value:
            break

        asn = str(asn_cell.value).strip()
        current_name = ws.cell(row=row_ptr, column=COL_APP_NAME).value
        current_sbg = ws.cell(row=row_ptr, column=COL_SBG).value
        current_owner = ws.cell(row=row_ptr, column=COL_OWNER).value
        current_status = ws.cell(row=row_ptr, column=COL_COMPLIANCE_STATUS).value

        if asn in master:
            matched_asns.add(asn)
            m = master[asn]
            display_name = current_name or m["name"]

            row_changes = []
            diff_and_apply(row_ptr, COL_APP_NAME, current_name, m["name"],
                            "Application Name", row_changes)
            diff_and_apply(row_ptr, COL_SBG, current_sbg, m["sbg"],
                            "SBG", row_changes)
            diff_and_apply(row_ptr, COL_OWNER, current_owner, m["owner"],
                            "Application Owner", row_changes)

            if row_changes:
                write_log_row(asn, display_name, " | ".join(row_changes))

            # Direct overwrite, no logging
            ws.cell(row=row_ptr, column=COL_FREQUENCY,
                    value=bucket_frequency(m["frequency"], m["frequency_unit"]))
            ws.cell(row=row_ptr, column=COL_COMPLIANCE_STATUS, value=m["internal_status"])
            if m["start_date"]:
                ws.cell(row=row_ptr, column=COL_DATE_MARKED, value=m["start_date"])

            tracker_rows_for_summary.append({
                "sbg": ws.cell(row=row_ptr, column=COL_SBG).value,
                "compliance_status": ws.cell(row=row_ptr, column=COL_COMPLIANCE_STATUS).value,
            })
        else:
            # In tracker but not in the latest Trackora dump - leave row untouched
            tracker_rows_for_summary.append({
                "sbg": current_sbg,
                "compliance_status": current_status,
            })

        row_ptr += 1

    last_row = row_ptr - 1

    # --- Append genuinely new ASNs ---
    new_asns = [a for a in master if a not in matched_asns]
    print(f"Appending {len(new_asns)} new application(s)...")
    append_row = last_row + 1
    for asn in new_asns:
        m = master[asn]
        ws.cell(row=append_row, column=COL_SECURETRACK, value=asn)
        ws.cell(row=append_row, column=COL_APP_NAME, value=m["name"])
        ws.cell(row=append_row, column=COL_SBG, value=m["sbg"])
        ws.cell(row=append_row, column=COL_OWNER, value=m["owner"])
        ws.cell(row=append_row, column=COL_FREQUENCY,
                value=bucket_frequency(m["frequency"], m["frequency_unit"]))
        ws.cell(row=append_row, column=COL_COMPLIANCE_STATUS, value=m["internal_status"])
        if m["start_date"]:
            ws.cell(row=append_row, column=COL_DATE_MARKED, value=m["start_date"])

        write_log_row(asn, m["name"], f"New application added (ASN {asn} not previously in tracker)")

        tracker_rows_for_summary.append({
            "sbg": m["sbg"],
            "compliance_status": m["internal_status"],
        })
        append_row += 1

    print("Recomputing COTS Summary and SBG breakdown tables...")
    update_summary_tables(ws, tracker_rows_for_summary)

    wb.save(args.output)

    print("-" * 50)
    print(f"Matched & updated: {len(matched_asns)}")
    print(f"New applications appended: {len(new_asns)}")
    print(f"Activity Log entries added this run: {changes_logged}")
    print(f"Saved to: {args.output}")
    print("-" * 50)
    print("NOTE: 'Compliance %' cells were written as raw fractions (e.g. 0.818),")
    print("relying on the cell's existing percentage number format to display")
    print("correctly. If it shows as '0.818' instead of '81.8%', just re-apply")
    print("percentage formatting to that cell once — the underlying value is correct.")


if __name__ == "__main__":
    main()