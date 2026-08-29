from app import app
from Categories.category_service import (
    get_template_folders,
    get_all_categories,
    )
from Bulk_Email.draft_bulk_sender import get_bulk_runs,send_bulk_from_draft
from flask import render_template, redirect, url_for, flash, request
import os,io,json
from graph_client import(
    get_draft_by_id,
    get_followup_drafts
)
from template_engine import extract_placeholders,render_dynamic
from openpyxl import load_workbook

@app.route("/bulk_email")
def bulk_email():
    folders=get_template_folders()
    drafts = []
    if folders:
        primary_folder_id=folders[0]
        drafts_response=get_followup_drafts(primary_folder_id)
        drafts=drafts_response.get("value",[])
    bulk_runs=get_bulk_runs()
    categories=get_all_categories()
    return render_template("bulk_email.html", drafts=drafts,bulk_runs=bulk_runs,categories=categories)

@app.route("/send_bulk_from_draft", methods=["POST"])
def send_bulk_from_draft_route():
 
    draft_id = request.form.get("draft_id")
    folders=get_template_folders()

    if not folders:
        flash("Folder settings not configured.")
        return redirect(url_for("bulk_email.html"))
    primary_folder_id = folders[0] or "Drafts"
    #secondary_folder_id = folders[1] or "Drafts"
    
    file = request.files.get("excel_file")
    #selected_draft_ids = request.form.get("selected_draft_ids", "")
    #selected_draft_ids = request.form.get("selected_draft_ids", "").strip()
 
 
    # ==========================
    # Send Mode Handling
    # ==========================
    bulk_category_name=request.form.get("bulk_category_name","").strip()
    workflow_type = request.form.get(
    "workflow_type",
    "generic"
    )
    server_mode = request.form.get(
    "server_mode",
    "hostname"
    )

    if not bulk_category_name:
        flash("Please select a category")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Follow-up Template Config
    # ==========================
    #followup_mode = request.form.get("followup_mode", "none")
    #default_followup_template_id = request.form.get("default_followup_template_id")
    #followup_sequence = request.form.getlist("followup_sequence")
 
    # ==========================
    # Basic Validation
    # ==========================
    if not draft_id:
        flash("Draft ID missing.")
        return redirect(url_for("bulk_email"))
 
    if not file:
        flash("No file uploaded.")
        return redirect(url_for("bulk_email"))

    # ==========================
    # Server Workflow Handling
    # ==========================
    if workflow_type == "server":
        try:
            from server_workflow.workflow_service import (
            prepare_server_workflow
            )
            # ==========================
            # Save uploaded file temporarily
            # ==========================
            temp_path = os.path.join(
                "temp_uploads",
                file.filename
                )
            os.makedirs(
                "temp_uploads",
                exist_ok=True
                )
            file.save(temp_path)
            # ==========================
            # Prepare workflow data
            # ==========================
            recipients = prepare_server_workflow(
                file_path=temp_path,
                mode=server_mode
                )
            # ==========================
            # Fetch Draft
            # ==========================
            draft = get_draft_by_id(
                draft_id,
                primary_folder_id
            )
            if not draft:
                flash("Draft not found.")
                return redirect(
                    url_for("bulk_email")
                )
            subject = draft.get("subject", "")
            body = draft.get("body", {}).get("content", "")
            # ==========================
            # Extract TO and CC templates
            # ==========================
            def extract_recipient_template(recipient_list):
                if not recipient_list:
                    return ""
                emails = []
                for r in recipient_list:
                    address = r.get(
                        "emailAddress",
                        {}
                    ).get("address")
                    if address:
                        emails.append(address)
                return ",".join(emails)
            to_template = extract_recipient_template(
                draft.get("toRecipients")
            )
            cc_template = extract_recipient_template(
                draft.get("ccRecipients")
            )
            # ==========================
            # Preview First Row
            # ==========================
            first = recipients[0]
            preview_subject = render_dynamic(
                subject,
                first
            )
            preview_body = render_dynamic(
                body,
                first
            )
            preview_to = render_dynamic(
                to_template,
                first
            )
            preview_cc = render_dynamic(
                cc_template,
                first
            )
            # ==========================
            # Render Preview
            # ==========================
            return render_template(
                "preview.html",
                draft_id=draft_id,
                folder_id=primary_folder_id,
                recipients=recipients,
                preview_subject=preview_subject,
                preview_body=preview_body,
                preview_to=preview_to,
                preview_cc=preview_cc,
                bulk_category_name=bulk_category_name
            )
 
        except Exception as e:
            flash(f"Server workflow failed: {str(e)}")
            return redirect(
                url_for("bulk_email")
            )
 
 
    # ==========================
    # Load workbook from memory
    # ==========================
    try:
        in_memory = io.BytesIO(file.read())
        wb = load_workbook(in_memory)
        sheet = wb.active
    except Exception:
        flash("Invalid Excel file.")
        return redirect(url_for("bulk_email"))
 
    rows = list(sheet.iter_rows(values_only=True))
 
    if not rows:
        flash("Excel file is empty.")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Header Processing
    # ==========================
    headers = [str(h).strip().lower() for h in rows[0] if h]
 
    if not headers:
        flash("Excel header row is empty.")
        return redirect(url_for("bulk_email"))
 
    recipients = []
 
    # ==========================
    # Process Rows
    # ==========================
    for i, row in enumerate(rows[1:], start=2):
 
        row_dict = {}
 
        for j in range(len(headers)):
            cell_value = row[j] if j < len(row) else ""
            row_dict[headers[j]] = (
                str(cell_value).strip() if cell_value is not None else ""
            )
 
        recipients.append(row_dict)
 
    if not recipients:
        flash("No data rows found.")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Fetch Draft
    # ==========================
    draft = get_draft_by_id(draft_id,primary_folder_id)
 
    if not draft:
        flash("Draft not found.")
        return redirect(url_for("bulk_email"))
 
    subject = draft.get("subject", "")
    body = draft.get("body", {}).get("content", "")
 
    # ==========================
    # Extract TO and CC templates
    # ==========================
    def extract_recipient_template(recipient_list):
        if not recipient_list:
            return ""
        emails = []
        for r in recipient_list:
            address = r.get("emailAddress", {}).get("address")
            if address:
                emails.append(address)
        return ",".join(emails)
 
    to_template = extract_recipient_template(draft.get("toRecipients"))
    cc_template = extract_recipient_template(draft.get("ccRecipients"))
 
    # ==========================
    # Placeholder Validation
    # ==========================
    full_template_text = subject + body + to_template + cc_template
    placeholders = extract_placeholders(full_template_text)
 
    missing = [p for p in placeholders if p.lower() not in headers]
 
    if missing:
        flash(f"Missing columns for placeholders: {', '.join(missing)}")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Preview First Row
    # ==========================
    first = recipients[0]
 
    try:
        preview_subject = render_dynamic(subject, first)
        preview_body = render_dynamic(body, first)
        preview_to = render_dynamic(to_template, first)
        preview_cc = render_dynamic(cc_template, first)
    except Exception as e:
        flash(str(e))
        return redirect(url_for("bulk_email"))

    # ==========================
    # Render Preview
    # ==========================
    return render_template(
        "preview.html",
        draft_id=draft_id,
        folder_id=primary_folder_id,
        recipients=recipients,
        preview_subject=preview_subject,
        preview_body=preview_body,
        preview_to=preview_to,
        preview_cc=preview_cc,
        bulk_category_name=bulk_category_name
    )
@app.route("/confirm_send", methods=["POST"])
def confirm_send():
 
    draft_id = request.form.get("draft_id")
    #folder_id = request.form.get("folder_id")
    folders=get_template_folders()

    if not folders:
        flash("Folder settings not configured.")
        return redirect(url_for("bulk_email"))
    primary_folder_id = folders[0] or "Drafts"

    recipients_json = request.form.get("recipients")
    bulk_category_name = request.form.get("bulk_category_name", "").strip()
 
    # ==========================
    # Basic Validation
    # ==========================
    if not draft_id:
        flash("Draft ID missing.")
        return redirect(url_for("bulk_email"))
 
    if not recipients_json:
        flash("No recipients data received.")
        return redirect(url_for("bulk_email"))
 
    if not bulk_category_name:
        flash("Please select a category.")
        return redirect(url_for("bulk_email"))
 
    try:
        recipients = json.loads(recipients_json)
    except json.JSONDecodeError:
        flash("Invalid recipients format.")
        return redirect(url_for("bulk_email"))
 
    if not isinstance(recipients, list) or len(recipients) == 0:
        flash("Recipients list is empty.")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Send Emails
    # ==========================
    try:
 
        print("Bulk category:", bulk_category_name)
        print("Recipients count:", len(recipients))
 
        results = send_bulk_from_draft(
            draft_id,
            primary_folder_id,
            recipients,
            bulk_category_name
        )
 
        # ==========================
        # Results Summary
        # ==========================
        success_count = len(results.get("success", []))
        failed_count = len(results.get("failed", []))
 
        flash(f"Bulk complete: {success_count} sent, {failed_count} failed.")
 
        return redirect(url_for("bulk_email"))
 
    except Exception as e:
        print("BULK ERROR:",str(e))
        raise e
        #flash(f"Error during bulk send: {str(e)}")
        #return redirect(url_for("bulk_email"))
