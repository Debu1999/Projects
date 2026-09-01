from flask import Blueprint, render_template, request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from db import get_connection
from Auth.routes_auth import login_required

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/reports/audit")
@login_required
def audit_report_page():
    workspace_id = request.args.get("workspace_id")
    workspace_name = None

    if workspace_id:
        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT workspace_name
                    FROM workspaces
                    WHERE id = %s AND user_id = %s
                """, (workspace_id, session.get("user_id")))
                row = cursor.fetchone()
                workspace_name = row[0] if row else None
        finally:
            conn.close()

    return render_template(
        "audit_report.html",
        workspace_id=workspace_id,
        workspace_name=workspace_name
    )


@reports_bp.route("/reports/audit/download")
@login_required
def audit_report_download():
    user_id = session.get("user_id")

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    workspace_id = request.args.get("workspace_id")

    workspace_name = None

    if workspace_id:
        query = """
            SELECT
                f.subject,
                f.category_name,
                f.started_at,
                f.attempt_count,
                f.last_client_reply_at,
                f.is_ignored_reply,
                f.last_followup_sent_at,
                f.status,
                f.last_reply_subject
            FROM followups f
            INNER JOIN workspace_conversations wc
                ON wc.conversation_id = f.conversation_id
                AND wc.user_id = f.user_id
            WHERE f.user_id = %s
              AND wc.workspace_id = %s
        """
        params = [user_id, workspace_id]
    else:
        query = """
            SELECT
                subject,
                category_name,
                started_at,
                attempt_count,
                last_client_reply_at,
                is_ignored_reply,
                last_followup_sent_at,
                status,
                last_reply_subject
            FROM followups
            WHERE user_id = %s
        """
        params = [user_id]

    date_col = "f.started_at" if workspace_id else "started_at"

    if start_date:
        query += f" AND {date_col} >= %s"
        params.append(start_date)

    if end_date:
        query += f" AND {date_col} <= %s"
        params.append(end_date + " 23:59:59")

    query += f" ORDER BY {date_col} ASC"

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()

            if workspace_id:
                cursor.execute("""
                    SELECT workspace_name FROM workspaces
                    WHERE id = %s AND user_id = %s
                """, (workspace_id, user_id))
                wrow = cursor.fetchone()
                workspace_name = wrow[0] if wrow else "Unknown Workspace"
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    generated_on = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    period_label = "All Time"
    if start_date or end_date:
        period_label = f"{start_date or 'Beginning'} to {end_date or 'Now'}"

    scope_label = f"Workspace: {workspace_name}" if workspace_id else "Scope: All Conversations"

    ws.merge_cells("A1:I1")
    ws["A1"] = f"SendTrix Audit Report — Generated On: {generated_on}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    ws["A2"] = f"{scope_label}   |   Reporting Period: {period_label}   |   Total Conversations: {len(rows)}"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.row_dimensions[2].height = 18

    headers = [
        "Subject", "Category", "Initial Mail Sent", "Followups Sent",
        "Client Replied", "Reply Ignored", "Last Client Reply Date",
        "Last Followup Sent", "Current Status",
    ]

    header_row_idx = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = header_row_idx + 1
    for r in rows:
        (subject, category, started_at, attempt_count,
         last_client_reply_at, is_ignored_reply,
         last_followup_sent_at, status, last_reply_subject) = r

        client_replied = "Yes" if last_client_reply_at else "No"
        reply_ignored = ("Yes" if is_ignored_reply else "No") if last_client_reply_at else "N/A"

        ws.cell(row=row_idx, column=1, value=subject or "—")
        ws.cell(row=row_idx, column=2, value=category or "—")
        ws.cell(row=row_idx, column=3, value=started_at.strftime("%Y-%m-%d %H:%M") if started_at else "—")
        ws.cell(row=row_idx, column=4, value=attempt_count or 0)
        ws.cell(row=row_idx, column=5, value=client_replied)
        ws.cell(row=row_idx, column=6, value=reply_ignored)
        ws.cell(row=row_idx, column=7, value=last_client_reply_at.strftime("%Y-%m-%d %H:%M") if last_client_reply_at else "—")
        ws.cell(row=row_idx, column=8, value=last_followup_sent_at.strftime("%Y-%m-%d %H:%M") if last_followup_sent_at else "—")
        ws.cell(row=row_idx, column=9, value=status or "—")

        if reply_ignored == "Yes":
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                )

        row_idx += 1

    widths = [35, 18, 18, 14, 13, 12, 20, 20, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = f"A{header_row_idx + 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = f"_{workspace_name.replace(' ', '_')}" if workspace_id and workspace_name else ""
    filename = f"SendTrix_Audit_Report{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )