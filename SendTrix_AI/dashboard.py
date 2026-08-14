from flask import Flask, render_template, redirect, url_for, flash, request,jsonify,send_file
import requests
import sqlite3
import os
from pprint import pprint
import csv
import pandas as pd
from werkzeug.utils import secure_filename
from excel_report import generate_comparison_excel
from graph_client import get_followup_drafts,get_messages_in_conversation,get_latest_message_in_conversation,get_user_availability,send_followup_reply_manual
from datetime import datetime, timedelta, timezone
from process import process,refresh_conversations
from db import init_db, save_settings, restart_followup,get_followups_by_status,get_connection,get_all_activity,clone_category_version
from test_sync import sync
from draft_service import get_drafts
from draft_bulk_sender import send_bulk_from_draft
from db import log_activity,get_activity,init_db,get_all_categories,get_category_details,get_latest_category_version
from auth import get_access_token
from draft_bulk_sender import ensure_outlook_category_exists
from db import get_template_folders,save_template_folders_settings,get_unread_replies,ignore_reply,create_new_version,insert_snapshot_bulk
from db import save_comment,run_comparison,insert_raw_snapshot_bulk,get_recommended_action,get_send_mail_flag,get_all_recommended_actions,get_action_mappings
from db import get_latest_external_responder,get_workspaces,create_workspace,add_conversation_to_workspace
from agent_service import rephrase_draft_body
from db import update_ai_draft_status,get_workspace_rows,move_conversation_to_workspace,remove_conversation_from_workspace
from agent_service import analyze_reply, clean_email_body
from db import get_ai_draft, save_ai_draft,insert_evidence_upload,update_ai_result,get_rows
from zoneinfo import ZoneInfo
from trackora_nl_query import natural_language_to_sql, run_safe_query
from trackora.evidence.extractor import extract
from trackora.evidence.extraction_worker import extract_version_and_timestamp, check_compliance_window
from trackora.evidence.orchestrator import process_evidence
from draft_bulk_sender import send_consolidated_from_draft

init_db()
app = Flask(__name__)
app.secret_key = "super_secret_key"
 
#BASE_DIR = os.path.dirname(os.path.abspath(__file__))
#DB_NAME = os.path.join(BASE_DIR, "followups.db")
UPLOAD_FOLDER="uploads/Evidence" 
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

 
def get_dashboard_counts():
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("SELECT COUNT(*) FROM followups WHERE status = 'ACTIVE'")
    active = cursor.fetchone()[0]
 
    cursor.execute("SELECT COUNT(*) FROM followups WHERE status = 'COMPLETED'")
    completed = cursor.fetchone()[0]
 
    cursor.execute("SELECT COUNT(*) FROM followups WHERE status = 'CLIENT_REPLY'")
    client_reply = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM followups WHERE status = 'MANUAL_PAUSED'")
    manual_pause = cursor.fetchone()[0]
 
    cursor.execute("SELECT COUNT(*) FROM followups")
    total = cursor.fetchone()[0]
 
    conn.close()
 
    return active, completed, client_reply,manual_pause, total
# ===============================
# Fetch Rows
# ===============================

@app.route("/trackora/upload_evidence", methods=["POST"])
def upload_evidence():
 
    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400
 
    file = request.files["file"]
    appser_number = request.form.get("appser_number")
 
    if file.filename == "":
        return jsonify({"success": False, "message": "No file selected"}), 400
 
    if not appser_number:
        return jsonify({"success": False, "message": "Application Number missing"}), 400
 
    filename = secure_filename(file.filename)
 
    app_folder = os.path.join(UPLOAD_FOLDER, appser_number)
    os.makedirs(app_folder, exist_ok=True)
 
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{filename}"
 
    file_path = os.path.join(app_folder, filename)
    file.save(file_path)
    #extracted_text=extract(file_path)
    result = process_evidence(file_path,appser_number)
    
    print("========== FINAL EVIDENCE ==========")
    
    pprint(result)
    print("====================================")
 
    uploaded_at = datetime.now().isoformat()
    file_type = os.path.splitext(filename)[1].lower()
    upload_id = None
    evidence_upload_id = insert_evidence_upload(
        upload_id=upload_id,
        appser_number=appser_number,
        file_name=filename,
        file_path=file_path,
        file_type=file_type,
        uploaded_at=uploaded_at
    )
    update_ai_result(evidence_upload_id,result)
 
    return jsonify({
        "success": True,
        "message": "Evidence uploaded successfully",
        "file_path": file_path
    })
 

@app.route("/nl_query", methods=["POST"])
def nl_query():
    print("=== NL QUERY ROUTE HIT ===", flush=True)
    question = request.form.get("question", "")

    result = natural_language_to_sql(question)
    print("Generated SQL:", result["sql"], flush=True)   # <-- this line was missing

    if not result["sql"]:
        return {"success": False, "error": result["explanation"]}

    query_result = run_safe_query(result["sql"], get_connection)

    return {
        "success": query_result["success"],
        "sql": result["sql"],
        "explanation": result["explanation"],
        "columns": query_result["columns"],
        "rows": query_result["rows"],
        "error": query_result["error"],
    }
import pandas as pd
import io
from flask import send_file
from trackora_nl_query import is_safe_select

@app.route("/nl_query/download", methods=["POST"])
def nl_query_download():
    sql = request.form.get("sql", "")

    if not is_safe_select(sql):
        return {"error": "Unsafe query, download blocked"}, 400

    conn = get_connection()
    df = pd.read_sql_query(sql, conn)
    conn.close()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="trackora_query_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
@app.route("/applications")
def applications():
    conn = get_connection()
    cursor = conn.cursor()
 
    # Get all uploads
    cursor.execute("""
        SELECT id, file_name, created_at
        FROM uploads
        ORDER BY created_at DESC
    """)
    rows = cursor.fetchall()
 
    # Get current master
    cursor.execute("""
        SELECT upload_id 
        FROM master_control 
        WHERE is_active = 1
        LIMIT 1
    """)
    row = cursor.fetchone()
 
    master_id = row[0] if row else None
 
    conn.close()
 
    uploads = []
    for r in rows:
        uploads.append({
            "id": r[0],
            "file_name": r[1],
            "created_at": r[2]
        })
 
    return render_template(
        "cots.html",
        uploads=uploads,
        master_exists = master_id is not None,
        master_id = master_id,
        title=""
    )

@app.route("/applications/demote-master", methods=["POST"])
def demote_master():
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        UPDATE master_control
        SET is_active = 0
        WHERE is_active = 1
    """)
 
    conn.commit()
    conn.close()
 
    return jsonify({"message": "Master removed successfully"})

@app.route("/applications/send-master-mails", methods=["POST"])
def send_master_mails():
 
    data = request.json
    items = data.get("items", [])
 
    if not items:
        return jsonify({"sent": 0, "failed": 0})
 
    from db import get_connection
    conn = get_connection()
    cursor = conn.cursor()
 
    # 🔹 Get active upload
    cursor.execute("SELECT upload_id FROM master_control WHERE is_active = 1")
    row = cursor.fetchone()
 
    if not row:
        return jsonify({"error": "No active master"})
 
    upload_id = row[0]
 
    # 🔹 Get draft folder (reuse existing bulk config)
    folders = get_template_folders()
    if not folders:
        return jsonify({"error": "Folder not configured"})
 
    primary_folder_id = folders[0]
    #consolidated_applications = []
 
    total_sent = 0
    total_failed = 0
 
 
    for item in items:
 
        asn = item.get("asn")
        draft_id = item.get("draft_id")
        category = item.get("category")

        cursor.execute("""
        SELECT send_status
        FROM application_analysis
        WHERE upload_id=? AND appser_number=?
        """, (upload_id, asn))
        status_row = cursor.fetchone()
        if status_row and status_row[0] in ("ACTIVE","SENDING"):
            continue
 
        # 🔹 Fetch full row
        cursor.execute("""
        SELECT
            s.appser_number,
            s.appser_name,
            s.appser_install_status,
            s.owner_name,
            s.tech_owner_name,
            s.current_installed_version,
            s.vendor_name,
            a.frequency,
            a.comments,
            a.internal_status
 
        FROM applications_snapshot s
        LEFT JOIN application_analysis a
        ON s.appser_number = a.appser_number
        AND s.upload_id = a.upload_id
 
        WHERE s.upload_id = ? AND s.appser_number = ?
        """, (upload_id, asn))
 
        row_data = cursor.fetchone()
 
        if not row_data:
            total_failed += 1
            continue
 
        # 🔹 Convert row → dict
        columns = [col[0] for col in cursor.description]
        raw_data = dict(zip(columns, row_data))
 
        # 🔹 Clean column names (so drafts are easy to write)
        mapping = {
            "appser_number": "asn",
            "appser_name": "name",
            "appser_install_status": "status",
            "owner_name": "owner",
            "tech_owner_name": "tech_owner",
            "current_installed_version": "version",
            "vendor_name": "vendor"
        }
 
        recipient = {}
 
        for key, value in raw_data.items():
            new_key = mapping.get(key, key)
            recipient[new_key] = value if value else ""
 
        # 🔹 Required for mail sending
        recipient["to"] = recipient.get("owner")
        recipient["cc"] = recipient.get("tech_owner")
 
        try:
            cursor.execute("""
            UPDATE application_analysis
            SET send_status='SENDING'
            WHERE upload_id=? AND appser_number=?
            """, (upload_id, asn))
            conn.commit()
            result = send_bulk_from_draft(
                draft_id,
                primary_folder_id,
                [recipient],   # same format as Excel
                category
            )
            successes=result.get("success",[])
            failures=result.get("failed",[])

            if successes:
                conversation_id=successes[0].get("conversation_id")
                
                cursor.execute("""
                UPDATE application_analysis
                SET send_status='ACTIVE',
                last_sent_at=?,
                conversation_id=?
                WHERE upload_id=? AND appser_number=?
                """, (
                    datetime.now(timezone.utc).isoformat(),
                    conversation_id,
                    upload_id,
                    asn
                ))
                conn.commit()
            elif failures:
                cursor.execute("""
                UPDATE application_analysis
                SET send_status='NOT_SENT'
                WHERE upload_id=? AND appser_number=?
                """, (upload_id, asn))
                conn.commit()
 
            total_sent += len(successes)
            total_failed += len(failures)
        except Exception as e:
            print("MASTER MAIL ERROR:", str(e))
            cursor.execute("""
            UPDATE application_analysis
            SET send_status='NOT_SENT'
            WHERE upload_id=? AND appser_number=?
            """, (upload_id, asn))
            conn.commit()
 
            total_failed += 1
 
    conn.close()
 
    return jsonify({
        "sent": total_sent,
        "failed": total_failed
    })
@app.route("/applications/send-master-consolidated", methods=["POST"])
def send_master_consolidated():
 
    data = request.json or {}
    items = data.get("items", [])
 
    if not items:
        return jsonify({
            "sent": 0,
            "failed": 0,
            "error": "No applications selected"
        }), 400
 
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    try:
 
        # --------------------------------
        # Get active master upload
        # --------------------------------
 
        cursor.execute("""
            SELECT upload_id
            FROM master_control
            WHERE is_active = 1
        """)
 
        row = cursor.fetchone()
 
        if not row:
            return jsonify({
                "sent": 0,
                "failed": 0,
                "error": "No active master"
            }), 400
 
        upload_id = row[0]
 
        # --------------------------------
        # Get draft folder
        # --------------------------------
 
        folders = get_template_folders()
 
        if not folders:
            return jsonify({
                "sent": 0,
                "failed": 0,
                "error": "Folder not configured"
            }), 400
 
        primary_folder_id = folders[0]
 
        # --------------------------------
        # Build application list
        # --------------------------------
 
        consolidated_applications = []
 
        for item in items:
 
            asn = item.get("asn")
 
            if not asn:
                continue
 
            cursor.execute("""
                SELECT
                    s.appser_number,
                    s.appser_name,
                    s.appser_install_status,
                    s.owner_name,
                    s.tech_owner_name,
                    s.current_installed_version,
                    s.vendor_name,
                    a.frequency,
                    a.comments,
                    a.internal_status
 
                FROM applications_snapshot s
 
                LEFT JOIN application_analysis a
                    ON s.appser_number = a.appser_number
                    AND s.upload_id = a.upload_id
 
                WHERE s.upload_id = ?
                AND s.appser_number = ?
            """, (upload_id, asn))
 
            row_data = cursor.fetchone()
 
            if not row_data:
                print("Application not found:", asn)
                continue
 
            columns = [
                col[0]
                for col in cursor.description
            ]
 
            raw_data = dict(
                zip(columns, row_data)
            )
 
            application = {
                "asn": raw_data.get("appser_number") or "",
                "name": raw_data.get("appser_name") or "",
                "status": raw_data.get("appser_install_status") or "",
                "owner": raw_data.get("owner_name") or "",
                "tech_owner": raw_data.get("tech_owner_name") or "",
                "version": raw_data.get("current_installed_version") or "",
                "vendor": raw_data.get("vendor_name") or "",
                "frequency": raw_data.get("frequency") or "",
                "comments": raw_data.get("comments") or "",
                "internal_status": raw_data.get("internal_status") or ""
            }
 
            consolidated_applications.append(application)
 
        print("====================================")
        print(
            "CONSOLIDATED APPLICATIONS:",
            len(consolidated_applications)
        )
        print(consolidated_applications)
        print("====================================")
 
        if not consolidated_applications:
            return jsonify({
                "sent": 0,
                "failed": 0,
                "error": "No valid applications found"
            }), 400
 
        # --------------------------------
        # Get selected draft/category
        # --------------------------------
 
        draft_id = items[0].get("draft_id")
        category_name = items[0].get("category")
 
        if not draft_id:
            return jsonify({
                "sent": 0,
                "failed": 0,
                "error": "No draft selected"
            }), 400
 
        # --------------------------------
        # Render consolidated email
        # --------------------------------
 
        result = send_consolidated_from_draft(
            draft_id=draft_id,
            folder_id=primary_folder_id,
            applications=consolidated_applications,
            category_name=category_name
        )
 
        print("====================================")
        print("CONSOLIDATED PREVIEW RESULT")
        print(result)
        print("====================================")
 
        return jsonify({
            "sent": 0,
            "failed": 0,
            "message": "Consolidated email rendered successfully",
            "preview": result.get("preview", {})
        })
 
    except Exception as e:
 
        print(
            "CONSOLIDATED MASTER ERROR:",
            str(e)
        )
 
        return jsonify({
            "sent": 0,
            "failed": len(items),
            "error": str(e)
        }), 500
 
    finally:
 
        conn.close()
 
@app.route("/toggle_auto_followup/<int:id>", methods=["POST"])
def toggle_auto_followup(id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT conversation_id, auto_followup_enabled FROM followups WHERE id = ?", (id,))
    conversation_id, current = cursor.fetchone()
    new_value = 0 if current else 1
    cursor.execute("UPDATE followups SET auto_followup_enabled = ? WHERE id = ?", (new_value, id))
    conn.commit()
    conn.close()
    log_activity(id, f"Auto-followup toggled {'ON' if new_value else 'OFF'} by user")
    return {"auto_followup_enabled": new_value}
@app.route("/approve_reply/<int:id>", methods=["POST"])
def approve_reply(id):
    final_response = request.form.get("final_response", "")

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT conversation_id, last_reply_message_id FROM followups WHERE id = ?", (id,))
    conversation_id, message_id = cursor.fetchone()
    conn.close()

    send_followup_reply_manual(message_id, final_response)
    update_ai_draft_status(conversation_id, "approved", new_body=final_response)
    log_activity(id, "AI reply approved and sent")

    flash(f"Reply sent for Followup ID {id}.")
    return redirect(url_for("dashboard"))


@app.route("/rephrase_reply/<int:id>", methods=["POST"])
def rephrase_reply(id):
    instruction = request.form.get("instruction", "")

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT conversation_id, last_reply_subject, last_reply_body, ai_draft_body
        FROM followups WHERE id = ?
    """, (id,))
    conversation_id, subject, body, previous_draft = cursor.fetchone()
    conn.close()

    clean_body = clean_email_body(body)
    new_draft = rephrase_draft_body(previous_draft, instruction, subject, clean_body)

    if new_draft:
        update_ai_draft_status(conversation_id, "pending", new_body=new_draft)
        return {"suggested_response": new_draft}
    else:
        return {"suggested_response": previous_draft, "error": "Rephrase failed, showing previous version"}


@app.route("/decline_reply/<int:id>", methods=["POST"])
def decline_reply(id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT conversation_id FROM followups WHERE id = ?", (id,))
    (conversation_id,) = cursor.fetchone()
    conn.close()

    update_ai_draft_status(conversation_id, "declined")
    log_activity(id, "AI reply declined by user")

    flash(f"AI reply declined for Followup ID {id}. Conversation remains paused.")
    return redirect(url_for("dashboard"))
 
@app.route("/applications/master-data")
def get_master_data():
    conn = get_connection()
    cursor = conn.cursor()
 
    # 🔹 Get active master
    cursor.execute("""
        SELECT upload_id FROM master_control
        WHERE is_active = 1
    """)
    row = cursor.fetchone()
 
    if not row:
        return jsonify([])
 
    upload_id = row[0]
    run_compliance_engine(upload_id)
 
    # 🔥 JOIN snapshot + analysis
    cursor.execute("""
        SELECT 
            s.appser_number,
            s.appser_name,
            s.appser_install_status,
            s.owner_name,
            s.tech_owner_name,
            s.current_installed_version,
            s.vendor_name,
 
            a.frequency,
            a.frequency_unit,
            a.review_start_date,
            a.next_due_date,
            a.comments,
            a.internal_status,
            a.send_mail,
            a.compliance_mode,
            a.remediation_due_date,
            a.exception_reason,
            a.vendor_status,
            a.recommended_action,
            a.action_status,
            a.send_status,

            m.draft_id,
            m.category,

            mt.meeting_status,
            mt.meeting_start,
            mt.meeting_link
 
        FROM applications_snapshot s
        LEFT JOIN application_analysis a
        ON s.appser_number = a.appser_number
        AND s.upload_id = a.upload_id

        LEFT JOIN application_mail_config m
        ON s.appser_number=m.appser_number
        AND s.upload_id=m.upload_id

        LEFT JOIN application_meetings mt
        ON mt.id = (
        SELECT MAX(id)
        FROM application_meetings x
        WHERE x.appser_number = s.appser_number
        AND x.upload_id = s.upload_id
        )
 
 
        WHERE s.upload_id = ?
    """, (upload_id,))
 
    rows = cursor.fetchall()
    conn.close()
 
    data = []
    for r in rows:
        data.append({
            "asn": r[0],
            "name": r[1],
            "status": r[2],
            "owner": r[3],
            "tech_owner": r[4],
            "version": r[5],
            "vendor": r[6],
            "upload_id":upload_id,
 
            "frequency": r[7] or "",
            "frequency_unit":r[8] or "",
            "review_start_date":r[9] or "",
            "next_due_date":r[10] or "",
            "comments": r[11] or "",
            "internal_status": r[12] or "",
            "send_mail": r[13] or 0,
            "compliance_mode":r[14] or "FREQUENCY",
            "remediation_due_date":r[15] or "",
            "exception_reason":r[16] or "",
            "vendor_status":r[17] or "",
            "recommended_action":r[18] or "",
            "action_status":r[19] or "",
            "send_status":r[20] or "",

            "draft_id":r[21] or "",
            "category":r[22] or "",

            "meeting_status":r[23] or "",
            "meeting_start":r[24] or "",
            "meeting_link":r[25] or ""
        })
    print(data[0])
 
    return jsonify(data)
@app.route("/applications/master-data/download", methods=["GET"])
def download_master_data():
    from db import get_connection
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    from flask import Response
    from trackora_report_narrative import compute_compliance_summary, generate_summary_narrative

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT upload_id FROM master_control WHERE is_active = 1
    """)
    row = cursor.fetchone()
    if not row:
        return jsonify({"error": "No active master set"}), 400
    upload_id = row[0]

    # Same JOIN as get_master_data()
    cursor.execute("""
        SELECT
            s.appser_number, s.appser_name, s.appser_install_status,
            s.owner_name, s.tech_owner_name, s.current_installed_version,
            s.vendor_name,
            a.compliance_mode, a.frequency, a.frequency_unit,
            a.review_start_date, a.remediation_due_date, a.next_due_date,
            a.exception_reason, a.comments, a.internal_status,
            a.vendor_status, a.recommended_action, a.action_status,
            a.send_status,
            mt.meeting_status, mt.meeting_start
        FROM applications_snapshot s
        LEFT JOIN application_analysis a
            ON s.appser_number = a.appser_number AND s.upload_id = a.upload_id
        LEFT JOIN application_meetings mt
            ON mt.id = (
                SELECT MAX(id) FROM application_meetings x
                WHERE x.appser_number = s.appser_number AND x.upload_id = s.upload_id
            )
        WHERE s.upload_id = ?
    """, (upload_id,))

    rows = cursor.fetchall()

    columns = [
        "ASN", "Name", "Status", "Owner", "Tech Owner", "Version", "Vendor",
        "Mode", "Frequency", "Frequency Unit", "Start Date", "Due Date",
        "Next Due Date", "Exception Reason", "Comments", "Internal Status",
        "Vendor Status", "Recommended Action", "Action Status", "Mail Status",
        "Meeting Status", "Meeting Start"
    ]
    df = pd.DataFrame(rows, columns=columns)

    # --- NEW: compute summary + AI narrative ---
    summary = compute_compliance_summary(upload_id, conn)
    narrative = generate_summary_narrative(summary)
    conn.close()

    # Build a small summary dataframe for the new sheet
    summary_rows = [
        ["Total Applications", summary["total"]],
        ["Compliant", summary["compliant"]],
        ["Non-Compliant", summary["non_compliant"]],
        ["", ""],
        ["Breakdown by SBG", ""],
    ]
    for sbg, counts in summary["sbg_breakdown"].items():
        summary_rows.append([sbg, f"{counts['compliant']} compliant / {counts['non_compliant']} non-compliant"])

    summary_rows.append(["", ""])
    summary_rows.append(["Breakdown by Compliance Mode", ""])
    for mode, counts in summary["mode_breakdown"].items():
        summary_rows.append([mode, f"{counts['compliant']} compliant / {counts['non_compliant']} non-compliant"])

    summary_rows.append(["", ""])
    summary_rows.append(["Breakdown by Frequency", ""])
    for freq, counts in summary["freq_breakdown"].items():
        summary_rows.append([freq, f"{counts['compliant']} compliant / {counts['non_compliant']} non-compliant"])

    summary_rows.append(["", ""])
    summary_rows.append(["AI Summary", narrative])

    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        df.to_excel(writer, index=False, sheet_name="Master Analysis")
    output.seek(0)

    filename = f"Trackora_Master_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
from datetime import datetime, timezone, timedelta

@app.route("/applications/save-mail-config", methods=["POST"])
def save_config():
    data = request.json
 
    upload_id = data["upload_id"]
    asn = data["asn"]
    draft = data.get("draft_id")
    category = data.get("category")
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    INSERT INTO application_mail_config (upload_id, appser_number, draft_id, category)
    VALUES (?, ?, ?, ?)
    ON CONFLICT(upload_id, appser_number)
    DO UPDATE SET draft_id = excluded.draft_id,
                  category = excluded.category
    """, (upload_id, asn, draft, category))
 
    conn.commit()
    conn.close()
 
    return jsonify({"status": "saved"})

 
@app.route("/applications/save-analysis", methods=["POST"])
def save_analysis():
    data = request.json
 
    upload_id = data.get("upload_id")
    appser_number = data.get("asn")
    frequency = data.get("frequency")
    
    comments = data.get("comments")
    internal_status = data.get("internal_status")
    send_mail = data.get("send_mail")
    vendor_status=data.get("vendor_status")
    compliance_mode = data.get(
        "compliance_mode",
        "FREQUENCY"
    )
    remediation_due_date = data.get(
        "remediation_due_date"
    )
    review_start_date=data.get("review_start_date")
    print("SAVE_DEBUG:",frequency,review_start_date,internal_status)
    frequency_unit=data.get("frequency_unit","days")
    exception_reason = data.get(
        "exception_reason"
    )
 
    conn = get_connection()
    cursor = conn.cursor()
 
    now = datetime.now(timezone.utc)
 
    # ==========================================
    # COMPLIANCE DATE LOGIC
    # ==========================================
    next_due_date = None
    last_reviewed_date = None
    # ------------------------------------------
    # FREQUENCY MODE
    # ------------------------------------------
    if (compliance_mode=="FREQUENCY" and frequency and internal_status == "Compliant"and review_start_date):
        print("Entered frequency block")
        try:
            freq_days = int(frequency)
        except:
            freq_days = 30
        if frequency_unit=="minutes":
            start_dt=now
        else:
            start_dt = now
            print("REVIEW_START_DATE =", repr(review_start_date))
            print("START_DT_TYPE =", type(start_dt))
            print("START_DT_VALUE =", repr(start_dt))
            print("REVIEW_START_DATE:",review_start_date,"START_DATE:",start_dt)
        last_reviewed_date = review_start_date
        if frequency_unit == "minutes":
            next_due_date = (
                start_dt + timedelta(minutes=freq_days)
            ).isoformat()
        else:
            next_due_date = (
                start_dt + timedelta(days=freq_days)
            ).isoformat()
            print("NEXT_DUE_DATE:",next_due_date)
    # ------------------------------------------
    # DATE BASED MODE
    # ------------------------------------------
    elif (compliance_mode=="DATE_BASED" and remediation_due_date):
        due_dt=datetime.fromisoformat(remediation_due_date)
        due_dt=due_dt.replace(hour=23,minute=59,second=59)
        next_due_date = due_dt.isoformat()
        print("DATE_MODE_DUE:",next_due_date)
        last_reviewed_date = now.isoformat()

    recommended_action=get_recommended_action(internal_status,vendor_status)
    send_mail=get_send_mail_flag(recommended_action)
    cursor.execute("""
    SELECT
    recommended_action,
    action_status
    FROM application_analysis
    WHERE upload_id = ?
    AND appser_number = ?
    """, (
        upload_id,
        appser_number
    ))
    existing = cursor.fetchone()
    action_status = "NONE"
    pending_actions = [
    "SEND_EVIDENCE_REQUEST",
    "REQUEST_FOR_EVIDENCE_WITH_TIMESTAMP",
    "WAIT_FOR_VENDOR_ETA",
    "WAIT_FOR_DUE_DATE"
    ]
    if recommended_action in pending_actions:
        if existing:
            old_recommended_action = existing[0]
            old_action_status = existing[1]
            if old_recommended_action == recommended_action:
                action_status = old_action_status or "PENDING"
            else:
                action_status = "PENDING"
        else:
            action_status = "PENDING"
 

 
    # UPSERT logic (UPDATED)
    cursor.execute("""
    INSERT INTO application_analysis (
        upload_id, appser_number,
        frequency,frequency_unit, comments, internal_status, send_mail,compliance_mode,remediation_due_date,exception_reason,
        last_reviewed_date, next_due_date, updated_at,vendor_status,recommended_action,action_status,review_start_date
    )
    VALUES (?, ?, ?, ?,?, ?, ?, ?, ?, ?,?,?,?,?,?,?,?)
    ON CONFLICT(upload_id, appser_number)
    DO UPDATE SET
        frequency=excluded.frequency,
        frequency_unit=excluded.frequency_unit,
        comments=excluded.comments,
        internal_status=excluded.internal_status,
        send_mail=excluded.send_mail,
        compliance_mode=excluded.compliance_mode,
        remediation_due_date=excluded.remediation_due_date,
        exception_reason=excluded.exception_reason,
        last_reviewed_date=COALESCE(excluded.last_reviewed_date, application_analysis.last_reviewed_date),
        next_due_date=COALESCE(excluded.next_due_date, application_analysis.next_due_date),
        updated_at=excluded.updated_at,
        vendor_status=excluded.vendor_status,
        recommended_action=excluded.recommended_action,
        action_status=excluded.action_status,
        review_start_date=excluded.review_start_date
    """, (
        upload_id, appser_number,
        frequency,frequency_unit, comments, internal_status, send_mail,
        compliance_mode,remediation_due_date,exception_reason,
        last_reviewed_date, next_due_date,
        now.isoformat(),vendor_status,recommended_action,action_status,review_start_date
    ))
 
    conn.commit()
    conn.close()
 
    return jsonify({
        "message": "Saved",
        "recommended_action":recommended_action,
        "send_mail":send_mail,
        "action_status":action_status
    })
@app.route("/rename-upload/<int:upload_id>", methods=["POST"])
def rename_upload(upload_id):
 
    data = request.json
    new_name = data.get("file_name","").strip()
    if not new_name:
        conn.close()
        return jsonify({
            "message": "File name cannot be empty"
        }), 400
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        SELECT file_name
        FROM uploads
        WHERE id = ?
    """, (upload_id,))
 
    row = cursor.fetchone()
 
    if not row:
        conn.close()
        return jsonify({
            "message": "Upload not found"
        }), 404
 
    old_name = row[0]
    old_ext = os.path.splitext(old_name)[1]
    if not new_name.lower().endswith(old_ext.lower()):
        new_name += old_ext
    cursor.execute("""
    UPDATE uploads
    SET file_name = ?
    WHERE id = ?
    """, (
        new_name,
        upload_id
    ))
    conn.commit()
    conn.close()
    return jsonify({
        "message":"File renamed successfully"
    })
@app.route("/delete-upload/<int:upload_id>", methods=["POST"])
def delete_upload(upload_id):
 
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT stored_name
    FROM uploads
    WHERE id = ?
    """, (upload_id,))
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return jsonify({
            "message": "Upload not found"
        }), 404
    stored_name = row[0]
    file_path = os.path.join("uploads", stored_name)
    if os.path.exists(file_path):
        os.remove(file_path)
    cursor.execute("""
    DELETE FROM master_control
    WHERE upload_id = ?
    """, (upload_id,))

    cursor.execute("""
    DELETE FROM applications_snapshot
    WHERE upload_id = ?
    """, (upload_id,))
    
    cursor.execute("""
    DELETE FROM applications_raw_data
    WHERE upload_id = ?
    """, (upload_id,))
    
    cursor.execute("""
    DELETE FROM uploads
    WHERE id = ?
    """, (upload_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        "message": "Upload deleted successfully"
    })
 
 
from datetime import datetime, timedelta, timezone
 
@app.route("/applications/suggest-slots/<asn>")
def suggest_slots(asn):
    duration=int(request.args.get("duration",30))
 
    now = datetime.now()

    latest_person=get_latest_external_responder(asn)
    print("LATEST_PERSON:",latest_person)

    availability=get_user_availability(latest_person["email"])
    print("AVAILABILITY:")
    print(availability)

    working_zone=ZoneInfo("Asia/Kolkata")

    schedule_items = (
    availability["value"][0]
    ["scheduleItems"]
    )
    print("SCHEDULE_ITEMS:")
    print(schedule_items)

    busy_slots = []
    for item in schedule_items:
        start = (datetime.fromisoformat(item["start"]["dateTime"]).replace(tzinfo=ZoneInfo("UTC")).astimezone(working_zone))
        end = (datetime.fromisoformat(item["end"]["dateTime"]).replace(tzinfo=ZoneInfo("UTC")).astimezone(working_zone))
        busy_slots.append({
            "start": start,
            "end": end,
            "status": item["status"]
        })
    print("BUSY SLOTS:", busy_slots)

    working_hours = (availability["value"][0]["workingHours"])
    print("WORKING_HOURS:")
    print(working_hours)
    working_days = [day.lower()for day in working_hours["daysOfWeek"]]
    
    start_hour = int(working_hours["startTime"].split(":")[0])
    end_hour = int(working_hours["endTime"].split(":")[0])
    slots = []
    for day in range(0, 4):
        current_day = (
            datetime.now(working_zone)
            + timedelta(days=day)
        )
        weekday_name = current_day.strftime("%A").lower()
        if weekday_name not in working_days:
            print(f"Skipping {weekday_name} (non-working day)")
            continue
        current = current_day.replace(
            hour=start_hour,
            minute=0,
            second=0,
            microsecond=0
        )
        if day == 0:
            now = datetime.now(working_zone)
            current = now.replace(second=0, microsecond=0)
            if current.minute == 0:
                pass
            elif current.minute <= 30:
                current = current.replace(minute=30)
            else:
                current = (
                    current.replace(minute=0)
                    + timedelta(hours=1)
                )
        work_end = current_day.replace(
            hour=end_hour,
            minute=0,
            second=0,
            microsecond=0
        )
        while current < work_end:
            slot = current
            slot_end = slot + timedelta(minutes=duration)
            is_busy = False
        
            slot_status="Available"
            for busy in busy_slots:
                busy_start=busy["start"]
                busy_end=busy["end"]
                
                if slot < busy_end and slot_end > busy_start:
                    slot_status=busy["status"]
                    if busy["status"] in ["busy","oof"]:
                        is_busy = True
                        break
            if not is_busy:
                slots.append({
                    "time":slot.isoformat(),
                    "status":slot_status
                })
            if len(slots)>=8:
                break
            current += timedelta(minutes=30)
    
    if not slots:
        print("No free slots found")
 
    #return jsonify(slots)
    return jsonify({
    "person": latest_person,
    "slots": slots,
    "workingHours": {
        "start": working_hours["startTime"],
        "end": working_hours["endTime"],
        "timezone": working_hours["timeZone"]["name"]
    }
    })

@app.route(
"/applications/create-meeting",
methods=["POST"]
)
def create_meeting():
 
    data = request.json
 
    asn = data.get("asn")
    start_time = data.get("start_time")
    duration = int(
        data.get("duration", 30)
    )
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    SELECT upload_id
    FROM master_control
    WHERE is_active = 1
    """)
 
    row = cursor.fetchone()
 
    if not row:
        return jsonify({
            "error":"No active master"
        })
 
    upload_id = row[0]
    cursor.execute("""
    SELECT conversation_id
    FROM application_analysis
    WHERE upload_id = ?
    AND appser_number = ?
    """, (
        upload_id,
        asn
    ))
    conv_row = cursor.fetchone()
    conversation_id = None
    if conv_row:
        conversation_id = conv_row[0]
    print("CONVERSATION_ID:", conversation_id)
    attendee_emails = set()
    if conversation_id:
        messages = get_messages_in_conversation(conversation_id)
        print("MESSAGE_COUNT:", len(messages))
        for msg in messages:
            sender = (
                msg.get("from", {})
                .get("emailAddress", {})
                .get("address", "")
                .lower()
                .strip()
            )
            if sender:
                attendee_emails.add(sender)
            for r in msg.get("toRecipients", []):
                email = (
                    r.get("emailAddress", {})
                    .get("address", "")
                    .lower()
                    .strip()
                )
                if email:
                    attendee_emails.add(email)
            for r in msg.get("ccRecipients", []):
                email = (
                    r.get("emailAddress", {})
                    .get("address", "")
                    .lower()
                    .strip()
                )
                if email:
                    attendee_emails.add(email)
    print("ATTENDEES_FOUND:", attendee_emails)
    attendees = []
    for email in attendee_emails:
        attendees.append({
            "emailAddress": {
                "address": email
            },
            "type": "required"
        })
 
    start_dt = datetime.fromisoformat(
        start_time
    )
 
    end_dt = start_dt + timedelta(
        minutes=duration
    )
    token = get_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    event_payload = {
        "subject": f"Trackora Compliance Review - {asn}",
        "start": {
            "dateTime": start_dt.isoformat(),
            "timeZone": "UTC"
        },
        "end": {
            "dateTime": end_dt.isoformat(),
            "timeZone": "UTC"
        },
        "attendees":attendees,
        "isOnlineMeeting": True,
        "onlineMeetingProvider":"teamsForBusiness"
    }
    print("ATTENDEE_COUNT:", len(attendees))
    print("ATTENDEES_PAYLOAD:", attendees)
    response = requests.post(
        "https://graph.microsoft.com/v1.0/me/events",
        headers=headers,
        json=event_payload
    )
    meeting_data = response.json()
    print("GRAPH_MEETING:", meeting_data)

    meeting_link = None
    if response.status_code in [200, 201]:
        meeting_link = (
            meeting_data.get("onlineMeeting", {}).get("joinUrl")
        )
        print("MEETING_LINK:", meeting_link)
    event_id = None
    if response.status_code in [200, 201]:
        meeting_link = (
            meeting_data.get("onlineMeeting", {}).get("joinUrl")
        )
        event_id = meeting_data.get("id")
    print("EVENT_ID:", event_id)

    latest_message = None
    if conversation_id:
        latest_message = get_latest_message_in_conversation(conversation_id)
    print("LATEST_MESSAGE:", latest_message)

    if latest_message and meeting_link:
        message_id = latest_message.get("id")
        reply_payload = {
            "comment": f"""Teams meeting has been scheduled.

            Join Meeting:{meeting_link}
            Meeting Time:{start_dt.strftime('%d-%b-%Y %H:%M UTC')}
            """
        }
        reply_response = requests.post(
            f"https://graph.microsoft.com/v1.0/me/messages/{message_id}/reply",
            headers=headers,
            json=reply_payload
        )
        print(
            "REPLY_STATUS:",
            reply_response.status_code
        )
        print(
            "REPLY_RESPONSE:",
            reply_response.text
        )
 
 
    cursor.execute("""
    SELECT id
    FROM application_meetings
    WHERE upload_id = ?
    AND appser_number = ?
    """, (
        upload_id,
        asn
    ))
    existing = cursor.fetchone()
    if existing:
        cursor.execute("""
        UPDATE application_meetings
        SET
        meeting_status = ?,
        meeting_start = ?,
        meeting_end = ?,
        meeting_link = ?,
        event_id=?,
        created_at = ?
        WHERE id = ?
        """, (
            "SCHEDULED",
            start_dt.isoformat(),
            end_dt.isoformat(),
            meeting_link,
            event_id,
            datetime.now().isoformat(),
            existing[0]
        ))
    else:
        cursor.execute("""
        INSERT INTO application_meetings (
        upload_id,
        appser_number,
        meeting_status,
        meeting_start,
        meeting_end,
        meeting_link,
        event_id,
        created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?,?)
        """, (
            upload_id,
            asn,
            "SCHEDULED",
            start_dt.isoformat(),
            end_dt.isoformat(),
            meeting_link,
            event_id,
            datetime.now().isoformat()
        ))
 
 
    conn.commit()
    conn.close()
 
    return jsonify({
      "message":"Meeting saved",
      "meeting_link":meeting_link
    })

@app.route(
    "/applications/send-action-mails",
    methods=["POST"]
)
def send_action_mails():
 
    conn = get_connection()
    cursor = conn.cursor()
    folders = get_template_folders()
    if not folders:
        return jsonify({
            "error": "Folder not configured"
        })
    primary_folder_id = folders[0]
 
    cursor.execute("""
    SELECT
    s.appser_number,
    s.appser_name,
    s.owner_name,
    s.tech_owner_name,
    a.recommended_action,
    a.action_status
    FROM application_analysis a
    JOIN applications_snapshot s
    ON a.upload_id=s.upload_id
    AND a.appser_number=s.appser_number
    WHERE a.action_status='PENDING'
    AND a.send_status='ACTIVE'
    """)
 
    rows = cursor.fetchall()
    from db import get_action_mapping
    sample = []
    for row in rows:
        recommended_action = row[4]
        mapping = get_action_mapping(
            recommended_action
        )
        if not mapping:
            continue
        draft_id = mapping["draft_id"]
        category = mapping["category_name"]
        recipient = {
            "asn": row[0],
            "name": row[1],
            "owner": row[2],
            "tech_owner": row[3],
            "to": row[2],
            "cc": row[3]
        }
        result = send_bulk_from_draft(
            draft_id,
            primary_folder_id,
            [recipient],
            category
        )
        sample.append({
            "action": recommended_action,
            "result": result
        })
    return jsonify({
        "count": len(rows),
        "sample": sample
    })
 
 
    conn.close()
 
    return jsonify({
        "count": len(rows)
    })
 
@app.route("/applications/action-mappings")
def action_mappings():
 
    existing = {
        row[0]: {
            "draft_id": row[1],
            "category_name": row[2]
        }
        for row in get_action_mappings()
    }
 
    result = []
 
    for action in get_all_recommended_actions():
 
        mapping = existing.get(action, {})
 
        result.append({
            "recommended_action": action,
            "draft_id": mapping.get("draft_id", ""),
            "category_name": mapping.get("category_name", "")
        })
 
    return jsonify(result)
@app.route(
    "/applications/save-action-mapping",
    methods=["POST"]
)
def save_action_mapping():
 
    data = request.json
 
    recommended_action = data.get(
        "recommended_action"
    )
 
    draft_id = data.get("draft_id")
    category_name = data.get("category_name")
 
    conn = get_connection()
    cursor = conn.cursor()
 
    now = datetime.now(timezone.utc)
 
    cursor.execute("""
    INSERT INTO action_template_mapping (
        recommended_action,
        draft_id,
        category_name,
        updated_at
    )
    VALUES (?, ?, ?, ?)
 
    ON CONFLICT(recommended_action)
    DO UPDATE SET
        draft_id=excluded.draft_id,
        category_name=excluded.category_name,
        updated_at=excluded.updated_at
    """, (
        recommended_action,
        draft_id,
        category_name,
        now.isoformat()
    ))
 
    conn.commit()
    conn.close()
 
    return jsonify({
        "message": "Mapping saved"
    })
 



@app.route("/applications/update-change-status", methods=["POST"])
def update_change_status():
 
    from db import get_connection
 
    data = request.json
 
    change_id = data.get("change_id")
    status = data.get("status")
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    UPDATE comparison_changes
    SET approval_status = ?
    WHERE id = ?
    """, (status, change_id))
 
    conn.commit()
    conn.close()
 
    return jsonify({
        "success": True
    })
 

@app.route("/applications/get-drafts")
def get_master_drafts():
    folders = get_template_folders()
 
    if not folders:
        return jsonify([])
 
    primary_folder_id = folders[0]
 
    drafts_response = get_followup_drafts(primary_folder_id)
    drafts = drafts_response.get("value", [])
 
    result = []
 
    for d in drafts:
        result.append({
            "id": d.get("id"),
            "subject": d.get("subject", "No Subject")
        })
 
    return jsonify(result)


'''@app.route("/applications/comparisons")
def list_comparisons():
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        SELECT 
            cl.id,
            cl.created_at,
            u1.file_name,
            u2.file_name
        FROM comparison_logs cl
        JOIN uploads u1 ON cl.from_version = u1.id
        JOIN uploads u2 ON cl.to_version = u2.id
        ORDER BY cl.id DESC
    """)
 
    rows = cursor.fetchall()
    conn.close()
 
    data = []
 
    for r in rows:
        data.append({
            "id": r[0],
            "name": f"Comp {r[0]}",
            "from_file": r[2],
            "to_file": r[3],
            "created_at": r[1]
        })
 
    return jsonify(data)'''
 
 

@app.route("/applications/comment", methods=["POST"])
def add_comment():
    data = request.json
 
    appser_number = data.get("appser_number")
    comment = data.get("comment")
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    SELECT id FROM master_versions
    WHERE is_active = 1
    LIMIT 1
    """)
    row = cursor.fetchone()
    conn.close()
 
    if not row:
        return jsonify({"message": "No active version"}), 400
 
    version_id = row[0]
 
    save_comment(appser_number, version_id, comment)
 
    return jsonify({"message": "Comment saved"})
    


@app.route("/get_reply/<int:id>")
def get_reply(id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT conversation_id, last_reply_subject, last_reply_body,
               last_client_reply_at,last_client_email
        FROM followups WHERE id = ?
    """, (id,))
    row = cursor.fetchone()
    conn.close()

    conversation_id, subject, body, last_reply_at,client_email = row

    saved = get_ai_draft(conversation_id)

    if saved and saved["analyzed_at"] and last_reply_at and last_reply_at <= saved["analyzed_at"]:
        print("Using saved AI analysis")
        analysis = saved
    else:
        print("Generating new AI analysis")
        clean_body = clean_email_body(body)
        result = analyze_reply(subject, clean_body,client_email)
        if result:
            save_ai_draft(conversation_id, result["draft_body"], result["reasoning"], result["classification"])
            analysis = result
        else:
            analysis = {"draft_body": "", "reasoning": "", "classification": "unclear"}

    return {
        "subject": subject,
        "body": body,
        "suggested_response": analysis["draft_body"],
        "summary": analysis["reasoning"],
        "classification": analysis["classification"],
    }
 
# ===============================
# Routes
# ===============================

from db import get_bulk_runs
@app.route("/category_versions/<category_name>")
def category_versions(category_name):
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        SELECT version
        FROM settings
        WHERE category_name = ?
        ORDER BY version
    """, (category_name,))
 
    rows = cursor.fetchall()
    conn.close()
 
    versions = [r[0] for r in rows]
 
    return {"versions": versions}
  

@app.route("/login")
def login_microsoft():
    try:
        token = get_access_token()
        flash("Microsoft login successful.")
    except Exception as e:
        flash(f"Login failed: {str(e)}")
 
    return redirect(url_for("dashboard"))
@app.route("/")
def dashboard():
    search=request.args.get("search","")
    rows=get_rows()
    unread_replies=get_unread_replies()
 
    active, completed, client_reply,manual_pause, total = get_dashboard_counts()
 
    print("Counts:", active, completed, client_reply,manual_pause, total)
 
    return render_template(
        "dashboard.html",
        rows=rows,
        active_count=active,
        completed_count=completed,
        client_reply_count=client_reply,
        manual_pause_count=manual_pause,
        total_count=total,
        unread_replies=unread_replies,
        search=search
    )
def carry_forward_mail_config(old_master_id, new_master_id):
    print("MAIL COPY")
    print("OLD =", old_master_id)
    print("NEW =", new_master_id)
 
    if not old_master_id:
        return
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    DELETE FROM application_mail_config
    WHERE upload_id = ?
    """, (new_master_id,))
 
    cursor.execute("""
    INSERT INTO application_mail_config
    (
        upload_id,
        appser_number,
        draft_id,
        category
    )
    SELECT
        ?,
        appser_number,
        draft_id,
        category
    FROM application_mail_config
    WHERE upload_id = ?
    """, (
        new_master_id,
        old_master_id
    ))
 
    conn.commit()
    conn.close()
def carry_forward_meetings(old_master_id, new_master_id):

    print("MEETING COPY")
    print("OLD =", old_master_id)
    print("NEW =", new_master_id)
 
 
    if not old_master_id:
        return
 
    conn = get_connection()
    cursor = conn.cursor()
 
    # Remove any existing meeting rows for the new upload
    cursor.execute("""
    DELETE FROM application_meetings
    WHERE upload_id = ?
    """, (new_master_id,))
 
    # Copy all meeting rows from old master
    cursor.execute("""
    INSERT INTO application_meetings
    (
        upload_id,
        appser_number,
        meeting_status,
        meeting_start,
        meeting_end,
        meeting_link,
        event_id,
        created_at
    )
    SELECT
        ?,
        appser_number,
        meeting_status,
        meeting_start,
        meeting_end,
        meeting_link,
        event_id,
        created_at
    FROM application_meetings
    WHERE upload_id = ?
    """, (
        new_master_id,
        old_master_id
    ))
 
    conn.commit()
    conn.close()
 
def initialize_or_carry_analysis(old_master_id, new_master_id):
    conn = get_connection()
    cursor = conn.cursor()

    if old_master_id is None:
        conn.close()
        return

    print("=" * 50)
    print("OLD MASTER:", old_master_id)
    print("NEW MASTER:", new_master_id)
    print("=" * 50)

    # Check whether this upload already has analysis
    cursor.execute("""
    SELECT COUNT(*)
    FROM application_analysis
    WHERE upload_id = ?
    """, (new_master_id,))
    existing_rows = cursor.fetchone()[0]
    
    if existing_rows > 0:
        print("Analysis already exists for upload:", new_master_id)
        conn.close()
        return
 
    # Get old analysis (if exists)
    old_data = {}
    cursor.execute("""
    SELECT *
    FROM application_analysis
    WHERE upload_id=?
    """,(old_master_id,))
    columns = [d[0] for d in cursor.description]
        
    for row in cursor.fetchall():
        data = dict(zip(columns, row))
        old_data[data["appser_number"]] = data
 
    # Get new snapshot ASNs
    cursor.execute("""
        SELECT appser_number FROM applications_snapshot
        WHERE upload_id = ?
    """, (new_master_id,))
    new_rows = [r[0] for r in cursor.fetchall()]
 
    for asn in new_rows:
        if asn in old_data:
            data=old_data[asn]
            # Remove old primary key
            data.pop("id", None)
            # Point to the new master
            data["upload_id"] = new_master_id
            # Fresh timestamp
            data["updated_at"] = datetime.now(timezone.utc).isoformat()
            # Start a new review cycle
            #data["last_reviewed_date"] = None
            #data["next_due_date"] = None
            #data["review_start_date"]
            print("COPYING ASN:",asn)
            print(data)
            cols = ",".join(data.keys())
            placeholders = ",".join(["?"] * len(data))
            cursor.execute(f"""
            INSERT INTO application_analysis
            ({cols})
            VALUES ({placeholders})
            """,
            tuple(data.values())
            )
        else:
            cursor.execute("""
            INSERT INTO application_analysis
            (
            upload_id,
            appser_number,
            updated_at
            )
            VALUES (?, ?, ?)
            """, (
                new_master_id,
                asn,
                datetime.now(timezone.utc).isoformat()
            ))
    cursor.execute("""
    SELECT COUNT(*)
    FROM application_analysis
    WHERE upload_id = ?
    """, (new_master_id,))
    print("Rows before commit:", cursor.fetchone()[0])
    conn.commit()
    conn.close()
@app.route("/upload", methods=["POST"])
def upload_application_file():
    import time
    import os
 
    file = request.files.get("file")

    if not file:
        return jsonify({"message": "No file uploaded"}), 400

    allowed = [".csv", ".xlsx"]
    if not any(file.filename.lower().endswith(ext)for ext in allowed):
        return jsonify({
            "message":"Only CSV and XLSX files are supported"
        }), 400
 
    UPLOAD_FOLDER = "uploads"
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
 
    original_name = file.filename
    unique_name = f"{int(time.time())}_{file.filename}"
 
    file_path = os.path.join(UPLOAD_FOLDER, unique_name)
 
    # ✅ Save permanently
    file.save(file_path)
 
    conn = get_connection()
    cursor = conn.cursor()
 
    now = datetime.now(timezone.utc).isoformat()
 
    # ✅ Save upload record
    cursor.execute("""
    INSERT INTO uploads (file_name, stored_name, created_at)
    VALUES (?, ?, ?)
    """, (original_name, unique_name, now))
 
    upload_id = cursor.lastrowid
    conn.commit()
    conn.close()
 
    # ✅ Insert snapshot data
    count = 0
    conn=get_connection()
    cursor=conn.cursor()
 
    if file.filename.lower().endswith(".csv"):
        with open(file_path,newline='',encoding='latin-1') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                if row.get("appser_number"):
                    insert_snapshot_bulk(cursor,upload_id, row)
                    insert_raw_snapshot_bulk(cursor,upload_id, row)
                    count += 1
    elif file.filename.lower().endswith(".xlsx"):
        df = pd.read_excel(file_path)
        df = df.fillna("")
        for _, row in df.iterrows():
            row_dict = row.to_dict()
            if row_dict.get("appser_number"):
                insert_snapshot_bulk(cursor,upload_id,row_dict)
                insert_raw_snapshot_bulk(cursor,upload_id,row_dict)
                count += 1
    conn.commit()
    cursor.execute("""
    SELECT COUNT(*)
    FROM applications_raw_data
    WHERE upload_id = ?
    """, (upload_id,))
    raw_count = cursor.fetchone()[0]
    
    cursor.execute("""
    SELECT COUNT(*)
    FROM applications_snapshot
    WHERE upload_id = ?
    """, (upload_id,))
    snapshot_count = cursor.fetchone()[0]
    
    print(
        "UPLOAD COMPLETE:",
        upload_id,
        "RAW:",
        raw_count,
        "SNAPSHOT:",
        snapshot_count
    )
    conn.close()

    return jsonify({
        "message": "Upload successful",
        "upload_id": upload_id,
        "records_inserted": count
    })
from datetime import datetime, timezone, timedelta
from flask import request, jsonify
 
@app.route("/applications/update-analysis", methods=["POST"])
def update_analysis():
    from db import get_connection
 
    data = request.json
 
    asn = data.get("asn")
    upload_id = data.get("upload_id")
 
    frequency = data.get("frequency")
    comments = data.get("comments")
    internal_status = data.get("internal_status")
    send_mail = data.get("send_mail")
 
    now = datetime.now(timezone.utc)
 
    last_reviewed_date = None
    next_due_date = None
 
    # 🔥 CORE LOGIC (your compliance engine starts here)
    if internal_status == "Compliant":
        try:
            freq_days = int(frequency)
        except:
            freq_days = 30
 
        last_reviewed_date = now.isoformat()
        next_due_date = (now + timedelta(days=freq_days)).isoformat()
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        UPDATE application_analysis
        SET frequency = ?,
            comments = ?,
            internal_status = ?,
            send_mail = ?,
            last_reviewed_date = COALESCE(?, last_reviewed_date),
            next_due_date = COALESCE(?, next_due_date),
            updated_at = ?
        WHERE upload_id = ? AND appser_number = ?
    """, (
        frequency,
        comments,
        internal_status,
        send_mail,
        last_reviewed_date,
        next_due_date,
        now.isoformat(),
        upload_id,
        asn
    ))
 
    conn.commit()
    conn.close()
 
    return jsonify({"message": "Saved"})
 
def run_compliance_engine(upload_id):
 
    conn = get_connection()
    cursor = conn.cursor()
 
    now = datetime.now(timezone.utc)
 
    cursor.execute("""
    SELECT appser_number, next_due_date, internal_status
    FROM application_analysis
    WHERE upload_id = ?
    """, (upload_id,))
 
    rows = cursor.fetchall()
 
    for r in rows:
        asn = r[0]
        next_due = r[1]
        status = r[2]
 
        if not next_due:
            continue
 
        try:
            next_due_dt = datetime.fromisoformat(next_due)
            if next_due_dt.tzinfo is None:
                next_due_dt=next_due_dt.replace(tzinfo=timezone.utc)
        except Exception as e:
            print("Data parse error:",e)
            continue
        print("CHECK:",asn,now,next_due_dt,status)
 
        # 🔥 CORRECT COMPARISON
        if now > next_due_dt and status == "Compliant":
            print("Expiring:",asn)
            cursor.execute("""
            UPDATE application_analysis
            SET internal_status = 'Non-compliant',
                send_mail = 1,
                updated_at = ?
            WHERE upload_id = ? AND appser_number = ?
            """, (now.isoformat(), upload_id, asn))
 
    conn.commit()
    conn.close()
@app.route("/debug/force-expire")
def force_expire():
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    UPDATE application_analysis
    SET next_due_date = '2000-01-01T00:00:00+00:00',
        internal_status = 'Compliant'
    WHERE internal_status = 'Compliant'
    """)
 
    conn.commit()
    conn.close()
 
    return "Forced expiry done"

 
 
 
@app.route("/applications/uploads", methods=["GET"])
def get_uploads():
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    SELECT u.id, u.file_name, u.created_at,
           CASE 
               WHEN m.upload_id IS NOT NULL THEN 1
               ELSE 0
           END as is_master
    FROM uploads u
    LEFT JOIN master_control m
    ON u.id = m.upload_id AND m.is_active = 1
    ORDER BY u.created_at DESC
    """)
 
    rows = cursor.fetchall()
    conn.close()
 
    uploads = []
    for r in rows:
        uploads.append({
            "id": r[0],
            "file_name": r[1],
            "created_at": r[2],
            "is_master": bool(r[3])
        })
 
    return jsonify(uploads)
@app.route("/applications/set-master/<int:upload_id>", methods=["POST"])
def set_master(upload_id):
    from db import get_connection
    from datetime import datetime, timezone
 
    conn = get_connection()
    cursor = conn.cursor()
 
    now = datetime.now(timezone.utc).isoformat()
 
    # 🔹 Get current active master (if any)
    cursor.execute("""
        SELECT upload_id FROM master_control
        WHERE is_active = 1
    """)
    row = cursor.fetchone()
    if row:
        old_master_id=row[0]
    else:
        old_master_id=upload_id
 
    # ❌ Remove existing master
    cursor.execute("""
        UPDATE master_control
        SET is_active = 0
        WHERE is_active = 1
    """)
 
    # ✅ Set new master
    cursor.execute("""
        INSERT INTO master_control (upload_id, is_active, created_at)
        VALUES (?, 1, ?)
    """, (upload_id, now))
 
    conn.commit()
    conn.close()
 
    # 🔥 NEW: Initialize or carry forward analysis
    #initialize_or_carry_analysis(old_master_id, upload_id)
    #carry_forward_mail_config(old_master_id,upload_id)
    #carry_forward_meetings(old_master_id,upload_id)

    run_compliance_engine(upload_id)
 
    return jsonify({"message": "Master set successfully"})

@app.route("/applications/compare/<int:upload_id>", methods=["POST"])
def compare_upload(upload_id):
    from db import get_connection
    #from comparison import run_comparison
 
    conn = get_connection()
    cursor = conn.cursor()
 
    # ✅ Get active master
    cursor.execute("""
    SELECT upload_id FROM master_control
    WHERE is_active = 1
    LIMIT 1
    """)
 
    row = cursor.fetchone()
    conn.close()
 
    if not row:
        return jsonify({"error": "No master selected"}), 400
 
    master_upload_id = row[0]
    print("MASTER_UPLOAD_ID:", master_upload_id)
    print("TARGET_UPLOAD_ID:", upload_id)
 
    comparison_id = run_comparison(master_upload_id, upload_id)
    
    return jsonify(
        {
            "message":"Comparison Completed",
            "comparison_id":comparison_id
        }
    )
@app.route("/applications/comparison-files")
def comparison_files():
 
    from db import get_comparison_files
 
    return jsonify(get_comparison_files())
@app.route("/applications/comparison-data/<int:comparison_id>", methods=["GET"])
def get_comparison_data(comparison_id):
 
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
    SELECT
        id,
        appser_number,
        field_name,
        old_value,
        new_value,
        change_type,
        approval_status
    FROM comparison_changes
    WHERE comparison_id = ?
    ORDER BY id DESC
    """, (comparison_id,))
 
    rows = cursor.fetchall()
 
    conn.close()
 
    data = []
 
    for r in rows:
 
        data.append({
            "id": r[0],
            "asn": r[1],
            "field": r[2],
            "old": r[3],
            "new": r[4],
            "type": r[5],
            "status": r[6]
        })
 
    return jsonify(data)
@app.route(
    "/applications/review-data/<int:comparison_id>",
    methods=["GET"]
)
def get_review_data(comparison_id):
 
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        SELECT
            appser_number,
            COUNT(*) as change_count,
            MAX(change_type) as change_type,
            MAX(approval_status) as approval_status
        FROM comparison_changes
        WHERE comparison_id = ?
        GROUP BY appser_number
        ORDER BY appser_number
    """, (comparison_id,))
 
    rows = cursor.fetchall()
    conn.close()
 
    data = []
 
    for r in rows:
        data.append({
            "asn": r[0],
            "change_count": r[1],
            "type": r[2],
            "status": r[3]
        })
 
    return jsonify(data)
@app.route(
    "/applications/asn-changes/<int:comparison_id>/<asn>",
    methods=["GET"]
)
def get_asn_changes(comparison_id, asn):
 
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        SELECT
        id,
        field_name,
        old_value,
        new_value,
        change_type,
        approval_status
        FROM comparison_changes
        WHERE comparison_id = ?
        AND appser_number = ?
        ORDER BY field_name
    """, (
        comparison_id,
        asn
    ))
 
    rows = cursor.fetchall()
 
    conn.close()
 
    data = []
 
    for r in rows:
        data.append({
            "id":r[0],
            "field": r[1],
            "old": r[2],
            "new": r[3],
            "type": r[4],
            "status": r[5]
        })
 
    return jsonify(data)
 
@app.route("/applications/apply-change/<int:change_id>", methods=["POST"])
def apply_change(change_id):
 
    from db import approve_change
 
    approve_change(change_id)
 
    return jsonify({
        "message": "Change applied"
    })
 
@app.route("/applications/ignore-change/<int:change_id>", methods=["POST"])
def ignore_change(change_id):
 
    from db import ignore_change_db
 
    ignore_change_db(change_id)
 
    return jsonify({
        "message": "Change ignored"
    })
@app.route(
    "/applications/apply-all/<int:comparison_id>",
    methods=["POST"]
)
def apply_all_changes(comparison_id):
 
    from db import apply_all_changes_db
 
    apply_all_changes_db(comparison_id)
 
    return jsonify({
        "message": "All changes applied"
    })
 
@app.route(
    "/applications/ignore-all/<int:comparison_id>",
    methods=["POST"]
)
def ignore_all_changes(comparison_id):
 
    from db import ignore_all_changes_db
 
    ignore_all_changes_db(comparison_id)
 
    return jsonify({
        "message": "All changes ignored"
    })
 
@app.route("/applications/download-final/<int:comparison_id>")
def download_final_excel(comparison_id):
 
    from excel_report import (
        generate_excel_from_upload
    )
 
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    # =====================================
    # GET CONSOLIDATED UPLOAD
    # =====================================
 
    cursor.execute("""
    SELECT id
    FROM uploads
    WHERE comparison_id = ?
    AND file_type = 'CONSOLIDATED'
    ORDER BY id DESC
    LIMIT 1
    """, (comparison_id,))
 
    row = cursor.fetchone()
 
    conn.close()
 
    if not row:
 
        return jsonify({
            "error": "No consolidated sheet found"
        }), 404
 
    upload_id = row[0]
 
    # =====================================
    # GENERATE EXCEL FROM CONSOLIDATED DATA
    # =====================================
 
    excel_path = generate_excel_from_upload(
        upload_id,
        comparison_id
    )
 
    return send_file(
        excel_path,
        as_attachment=True
    )
 
@app.route(
    "/applications/generate-consolidated/<int:comparison_id>",
    methods=["POST"]
)
def generate_consolidated(comparison_id):
 
    from excel_report import (
        create_upload_from_comparison
    )
 
    from db import get_connection
 
    conn = get_connection()
    cursor = conn.cursor()
 
    # =====================================
    # GET COMPARISON DETAILS
    # =====================================
 
    cursor.execute("""
    SELECT
        from_upload_id,
        to_upload_id
    FROM comparison_logs
    WHERE id = ?
    """, (comparison_id,))
 
    row = cursor.fetchone()
 
    conn.close()
 
    if not row:
        return jsonify({
            "error": "Comparison not found"
        }), 404
 
    master_upload_id = row[0]
    target_upload_id = row[1]
 
    # =====================================
    # CREATE CONSOLIDATED UPLOAD
    # =====================================
 
    new_upload_id = (
        create_upload_from_comparison(
            comparison_id,
            master_upload_id,
            target_upload_id
        )
    )
    initialize_or_carry_analysis(master_upload_id,new_upload_id)
    carry_forward_mail_config(master_upload_id,new_upload_id)
    carry_forward_meetings(master_upload_id,new_upload_id)
 
    return jsonify({
        "message": "Consolidated sheet generated",
        "upload_id": new_upload_id
    })
 
 
@app.route("/bulk_email")
def bulk_email():
    folders=get_template_folders()
    drafts = []
    if folders:
        primary_folder_id=folders[0]
        drafts_response=get_followup_drafts(primary_folder_id)
        drafts=drafts_response.get("value",[])
    bulk_runs=get_bulk_runs()
    categories=get_all_categories()
    return render_template("bulk_email.html", drafts=drafts,bulk_runs=bulk_runs,categories=categories)

from template_engine import extract_placeholders
from validators import is_valid_email
from draft_service import get_draft_by_id

import json
from draft_bulk_sender import send_bulk_from_draft

from openpyxl import load_workbook
from template_engine import extract_placeholders, render_dynamic
from validators import is_valid_email
from draft_service import get_draft_by_id
import io
 
@app.route("/send_bulk_from_draft", methods=["POST"])
def send_bulk_from_draft_route():
 
    draft_id = request.form.get("draft_id")
    folders=get_template_folders()

    if not folders:
        flash("Folder settings not configured.")
        return redirect(url_for("bulk_email.html"))
    primary_folder_id = folders[0] or "Drafts"
    #secondary_folder_id = folders[1] or "Drafts"
    
    file = request.files.get("excel_file")
    #selected_draft_ids = request.form.get("selected_draft_ids", "")
    #selected_draft_ids = request.form.get("selected_draft_ids", "").strip()
 
 
    # ==========================
    # Send Mode Handling
    # ==========================
    bulk_category_name=request.form.get("bulk_category_name","").strip()
    workflow_type = request.form.get(
    "workflow_type",
    "generic"
    )
    server_mode = request.form.get(
    "server_mode",
    "hostname"
    )

    if not bulk_category_name:
        flash("Please select a category")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Follow-up Template Config
    # ==========================
    #followup_mode = request.form.get("followup_mode", "none")
    #default_followup_template_id = request.form.get("default_followup_template_id")
    #followup_sequence = request.form.getlist("followup_sequence")
 
    # ==========================
    # Basic Validation
    # ==========================
    if not draft_id:
        flash("Draft ID missing.")
        return redirect(url_for("bulk_email"))
 
    if not file:
        flash("No file uploaded.")
        return redirect(url_for("bulk_email"))

    # ==========================
    # Server Workflow Handling
    # ==========================
    if workflow_type == "server":
        try:
            from server_workflow.workflow_service import (
            prepare_server_workflow
            )
            # ==========================
            # Save uploaded file temporarily
            # ==========================
            temp_path = os.path.join(
                "temp_uploads",
                file.filename
                )
            os.makedirs(
                "temp_uploads",
                exist_ok=True
                )
            file.save(temp_path)
            # ==========================
            # Prepare workflow data
            # ==========================
            recipients = prepare_server_workflow(
                file_path=temp_path,
                mode=server_mode
                )
            # ==========================
            # Fetch Draft
            # ==========================
            draft = get_draft_by_id(
                draft_id,
                primary_folder_id
            )
            if not draft:
                flash("Draft not found.")
                return redirect(
                    url_for("bulk_email")
                )
            subject = draft.get("subject", "")
            body = draft.get("body", {}).get("content", "")
            # ==========================
            # Extract TO and CC templates
            # ==========================
            def extract_recipient_template(recipient_list):
                if not recipient_list:
                    return ""
                emails = []
                for r in recipient_list:
                    address = r.get(
                        "emailAddress",
                        {}
                    ).get("address")
                    if address:
                        emails.append(address)
                return ",".join(emails)
            to_template = extract_recipient_template(
                draft.get("toRecipients")
            )
            cc_template = extract_recipient_template(
                draft.get("ccRecipients")
            )
            # ==========================
            # Preview First Row
            # ==========================
            first = recipients[0]
            preview_subject = render_dynamic(
                subject,
                first
            )
            preview_body = render_dynamic(
                body,
                first
            )
            preview_to = render_dynamic(
                to_template,
                first
            )
            preview_cc = render_dynamic(
                cc_template,
                first
            )
            # ==========================
            # Render Preview
            # ==========================
            return render_template(
                "preview.html",
                draft_id=draft_id,
                folder_id=primary_folder_id,
                recipients=recipients,
                preview_subject=preview_subject,
                preview_body=preview_body,
                preview_to=preview_to,
                preview_cc=preview_cc,
                bulk_category_name=bulk_category_name
            )
 
        except Exception as e:
            flash(f"Server workflow failed: {str(e)}")
            return redirect(
                url_for("bulk_email")
            )
 
 
    # ==========================
    # Load workbook from memory
    # ==========================
    try:
        in_memory = io.BytesIO(file.read())
        wb = load_workbook(in_memory)
        sheet = wb.active
    except Exception:
        flash("Invalid Excel file.")
        return redirect(url_for("bulk_email"))
 
    rows = list(sheet.iter_rows(values_only=True))
 
    if not rows:
        flash("Excel file is empty.")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Header Processing
    # ==========================
    headers = [str(h).strip().lower() for h in rows[0] if h]
 
    if not headers:
        flash("Excel header row is empty.")
        return redirect(url_for("bulk_email"))
 
    recipients = []
 
    # ==========================
    # Process Rows
    # ==========================
    for i, row in enumerate(rows[1:], start=2):
 
        row_dict = {}
 
        for j in range(len(headers)):
            cell_value = row[j] if j < len(row) else ""
            row_dict[headers[j]] = (
                str(cell_value).strip() if cell_value is not None else ""
            )
 
        recipients.append(row_dict)
 
    if not recipients:
        flash("No data rows found.")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Fetch Draft
    # ==========================
    draft = get_draft_by_id(draft_id,primary_folder_id)
 
    if not draft:
        flash("Draft not found.")
        return redirect(url_for("bulk_email"))
 
    subject = draft.get("subject", "")
    body = draft.get("body", {}).get("content", "")
 
    # ==========================
    # Extract TO and CC templates
    # ==========================
    def extract_recipient_template(recipient_list):
        if not recipient_list:
            return ""
        emails = []
        for r in recipient_list:
            address = r.get("emailAddress", {}).get("address")
            if address:
                emails.append(address)
        return ",".join(emails)
 
    to_template = extract_recipient_template(draft.get("toRecipients"))
    cc_template = extract_recipient_template(draft.get("ccRecipients"))
 
    # ==========================
    # Placeholder Validation
    # ==========================
    full_template_text = subject + body + to_template + cc_template
    placeholders = extract_placeholders(full_template_text)
 
    missing = [p for p in placeholders if p.lower() not in headers]
 
    if missing:
        flash(f"Missing columns for placeholders: {', '.join(missing)}")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Preview First Row
    # ==========================
    first = recipients[0]
 
    try:
        preview_subject = render_dynamic(subject, first)
        preview_body = render_dynamic(body, first)
        preview_to = render_dynamic(to_template, first)
        preview_cc = render_dynamic(cc_template, first)
    except Exception as e:
        flash(str(e))
        return redirect(url_for("bulk_email"))

    # ==========================
    # Render Preview
    # ==========================
    return render_template(
        "preview.html",
        draft_id=draft_id,
        folder_id=primary_folder_id,
        recipients=recipients,
        preview_subject=preview_subject,
        preview_body=preview_body,
        preview_to=preview_to,
        preview_cc=preview_cc,
        bulk_category_name=bulk_category_name
    )

from graph_client import get_outlook_drafts,get_draft_content,move_draft,create_draft_child_folder,get_draft_child_folders

from flask import jsonify
 
 
@app.route("/template_preview/<message_id>")
def template_preview(message_id):
 
    draft = get_draft_content(message_id)
 
    # Extract recipients
    to_list = [
        r.get("emailAddress", {}).get("address")
        for r in draft.get("toRecipients", [])
    ]
 
    cc_list = [
        r.get("emailAddress", {}).get("address")
        for r in draft.get("ccRecipients", [])
    ]
 
    return jsonify({
        "subject": draft.get("subject"),
        "body": draft.get("body", {}).get("content", ""),
        "to": to_list,
        "cc": cc_list
    })

@app.route("/move_template", methods=["POST"])
def move_template():
 
    message_id = request.json.get("message_id")
    folder_id = request.json.get("folder_id")
 
    move_draft(message_id, folder_id)
 
    return {"status": "success"}

@app.route("/create_template_folder", methods=["POST"])
def create_template_folder():
 
    folder_name = request.json.get("folder_name")
 
    create_draft_child_folder(folder_name)
 
    return {"status":"success"}

@app.route("/templates")
def templates():
 
    drafts_by_folder = get_outlook_drafts()
 
    folder_data = get_draft_child_folders()
    folders = folder_data.get("value", [])
 
    folders.insert(0,{
        "id":"Drafts",
        "displayName":"Drafts"
    })
 
    # NEW: fetch saved template folder settings
    folder_settings = get_template_folders()
 
    primary_folder_id = None
    secondary_folder_id = None
 
    if folder_settings:
        primary_folder_id = folder_settings[0]
        secondary_folder_id = folder_settings[1]
 
    return render_template(
        "templates_page.html",
        drafts_by_folder=drafts_by_folder,
        folders=folders,
        primary_folder_id=primary_folder_id,
        secondary_folder_id=secondary_folder_id
    )

@app.route("/categories")
def categories():
    return render_template("categories.html")

@app.route("/api/categories")
def categories_api():
    return jsonify(get_all_categories())

@app.route("/category/<name>/<int:version>")
def category_details(name,version):
 
    data = get_category_details(name,version)
 
    if not data:
        return jsonify({"error": "Category not found"}), 404
 
    return jsonify(data)

@app.route("/clone_category", methods=["POST"])
def clone_category():
 
    data = request.get_json()
 
    if not data:
        return jsonify({"error": "No JSON received"}), 400
 
    category_name = data.get("category_name")
    version = data.get("version")
 
    if not category_name or not version:
        return jsonify({"error": "Invalid request"}), 400
 
    try:
 
        new_version = clone_category_version(category_name, int(version))
 
        return jsonify({
            "success": True,
            "new_version": new_version
        })
 
    except Exception as e:
        print("CLONE ERROR:", e)
        return jsonify({"error": str(e)}), 500
 
@app.route("/fetch_drafts")
def fetch_drafts():
    print("Fetch Drafts Called")
 
    folders = get_template_folders()
    print("Folders from DB:",folders)
 
    if not folders:
        print("No folders found in DB")
        return {"value": []}

    primary_folder_id,secondary_folder_id=folders
    print("Folders from DB:",folders)
    print("Primary_Folder:",primary_folder_id)
    print("Using Folder:",secondary_folder_id)
 
    drafts_response = get_followup_drafts(secondary_folder_id)
    print("Graph Response:",drafts_response)
 
    return drafts_response

@app.route("/save_template_folders", methods=["POST"])
def save_template_folders():
    data = request.json
    initiate_id = data.get("primary_folder_id")
    followup_id = data.get("secondary_folder_id")
    save_template_folders_settings(initiate_id, followup_id)
    return {"status":"success"}

@app.route("/confirm_send", methods=["POST"])
def confirm_send():
 
    draft_id = request.form.get("draft_id")
    #folder_id = request.form.get("folder_id")
    folders=get_template_folders()

    if not folders:
        flash("Folder settings not configured.")
        return redirect(url_for("bulk_email"))
    primary_folder_id = folders[0] or "Drafts"

    recipients_json = request.form.get("recipients")
    bulk_category_name = request.form.get("bulk_category_name", "").strip()
 
    # ==========================
    # Basic Validation
    # ==========================
    if not draft_id:
        flash("Draft ID missing.")
        return redirect(url_for("bulk_email"))
 
    if not recipients_json:
        flash("No recipients data received.")
        return redirect(url_for("bulk_email"))
 
    if not bulk_category_name:
        flash("Please select a category.")
        return redirect(url_for("bulk_email"))
 
    try:
        recipients = json.loads(recipients_json)
    except json.JSONDecodeError:
        flash("Invalid recipients format.")
        return redirect(url_for("bulk_email"))
 
    if not isinstance(recipients, list) or len(recipients) == 0:
        flash("Recipients list is empty.")
        return redirect(url_for("bulk_email"))
 
    # ==========================
    # Send Emails
    # ==========================
    try:
 
        print("Bulk category:", bulk_category_name)
        print("Recipients count:", len(recipients))
 
        results = send_bulk_from_draft(
            draft_id,
            primary_folder_id,
            recipients,
            bulk_category_name
        )
 
        # ==========================
        # Results Summary
        # ==========================
        success_count = len(results.get("success", []))
        failed_count = len(results.get("failed", []))
 
        flash(f"Bulk complete: {success_count} sent, {failed_count} failed.")
 
        return redirect(url_for("bulk_email"))
 
    except Exception as e:
        print("BULK ERROR:",str(e))
        raise e
        #flash(f"Error during bulk send: {str(e)}")
        #return redirect(url_for("bulk_email"))



@app.route("/update_category", methods=["POST"])
def update_category():
 
    data = request.get_json()
 
    category_name = data.get("category_name")
    max_attempts = int(data.get("max_attempts"))
    interval_minutes = int(data.get("interval_minutes"))
    mode = data.get("mode")
    followup_text = data.get("followup_text", "")
    restart_sequences = data.get("restart_sequences", False)
 
    templates = data.get("templates", [])
 
    selected_draft_ids = ""
 
    if templates:
        selected_draft_ids = ",".join(
            [t.get("id") for t in templates if t.get("id")]
        )
 
    # 🔹 create new version using db.py
    new_version = save_settings(
        category_name,
        followup_text,
        max_attempts,
        interval_minutes,
        mode,
        selected_draft_ids
    )
 
    # 🔹 restart active sequences if requested
    if restart_sequences:
 
        conn = get_connection()
        cursor = conn.cursor()
 
        cursor.execute("""
        UPDATE followups
        SET attempt_count = 0,
            next_followup_at = CURRENT_TIMESTAMP,
            category_version = ?
        WHERE category_name = ?
        AND status = 'ACTIVE'
        """, (
            new_version,
            category_name
        ))
 
        conn.commit()
        conn.close()
 
    return jsonify({
        "status": "ok",
        "new_version": new_version
    })
 
 
@app.route("/activity/<int:id>")
def activity(id):
    conn =get_connection()
    cursor = conn.cursor()
 
    cursor.execute("SELECT conversation_id FROM followups WHERE id=?", (id,))
    row = cursor.fetchone()
    conn.close()
 
    if not row:
        flash("Conversation not found.")
        return redirect(url_for("dashboard"))
 
    conversation_id = row[0]
    logs = get_activity(conversation_id)
 
    return render_template("activity.html", logs=logs) 
@app.route("/activity")
def activity_page():
    logs = get_all_activity()
 
    formatted_logs = []
    for log in logs:
        subject, action, created_at = log
        formatted_logs.append(
            (subject, action, to_ist(created_at))
        )
 
    return render_template("activity.html", logs=formatted_logs)

@app.route("/run")
def run_process():
    try:
        get_access_token()
    except:
        flash("Please login to Microsoft first.")
        return redirect(url_for("login_microsoft"))
 
    process()
    flash("Followup process executed successfully.")
    return redirect(url_for("dashboard"))
@app.route("/workspace/<int:workspace_id>/run")
def run_workspace_process(workspace_id):
 
    try:
        process(workspace_id=workspace_id)
 
        flash("Workspace followups processed successfully.")
 
    except Exception as e:
        print("Workspace process error:", str(e))
        flash("Error processing workspace followups.")
 
    return redirect(
        url_for(
            "workspace_detail",
            workspace_id=workspace_id
        )
    )
 
@app.route("/refresh")
def refresh_process():
 
    try:
        get_access_token()
    except:
        flash("Please login to Microsoft first.")
        return redirect(url_for("login_microsoft"))
 
    refresh_conversations()
 
    flash("Conversations refreshed successfully.")
    return redirect(url_for("dashboard"))
@app.route("/workspace/<int:workspace_id>/refresh")
def refresh_workspace(workspace_id):
 
    try:
 
        refresh_conversations(
            workspace_id=workspace_id
        )
 
        flash(
            "Workspace conversations refreshed successfully."
        )
 
    except Exception as e:
 
        print(
            "Workspace refresh error:",
            str(e)
        )
 
        flash(
            "Error refreshing workspace conversations."
        )
 
    return redirect(
        url_for(
            "workspace_detail",
            workspace_id=workspace_id
        )
    )

@app.route("/sync")
def run_sync():
    sync()
    flash("Mailbox sync completed successfully.")
    return redirect(url_for("dashboard"))
 
 
# ===============================
# Restart (SAFE VERSION)
# ===============================
@app.route("/restart/<int:id>")
def restart(id):

    version=request.args.get("version")
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        SELECT conversation_id
        FROM followups
        WHERE id = ?
    """, (id,))
    row = cursor.fetchone()
    conn.close()
 
    if row:
        conversation_id = row[0]
        if version:
            restart_followup(conversation_id,int(version))
        else:
            restart_followup(conversation_id)
            flash(f"Followup ID {id} restarted successfully.")
 
    return redirect(url_for("dashboard"))
@app.route("/workspaces")
def workspaces_page():
 
    workspaces = get_workspaces()
 
    return render_template(
        "workspaces.html",
        title="",
        workspaces=workspaces
    )
@app.route("/workspace/<int:workspace_id>")
def workspace_detail(workspace_id):
 
    try:
 
        conn = get_connection()
        cursor = conn.cursor()
 
        cursor.execute("""
            SELECT
                id,
                workspace_name,
                description,
                workspace_type,
                status
            FROM workspaces
            WHERE id = ?
            AND status = 'ACTIVE'
        """, (workspace_id,))
 
        workspace_row = cursor.fetchone()
 
        conn.close()
 
        if not workspace_row:
            return "Workspace not found.", 404
 
        workspace = {
            "id": workspace_row[0],
            "workspace_name": workspace_row[1],
            "description": workspace_row[2],
            "workspace_type": workspace_row[3],
            "status": workspace_row[4]
        }
 
        # Get ONLY conversations belonging to this workspace
        rows = get_workspace_rows(workspace_id)
 
        # Calculate workspace-specific counts
        active_count = sum(
            1 for row in rows
            if row[3] == "ACTIVE"
        )
 
        completed_count = sum(
            1 for row in rows
            if row[3] == "COMPLETED"
        )
 
        client_reply_count = sum(
            1 for row in rows
            if row[3] == "CLIENT_REPLY"
        )
 
        manual_pause_count = sum(
            1 for row in rows
            if row[3] == "MANUAL_PAUSED"
        )
 
        total_count = len(rows)
 
        return render_template(
            "workspace_detail.html",
            workspace_id=workspace_id,
 
            title=workspace["workspace_name"],
 
            workspace=workspace,
 
            rows=rows,
 
            active_count=active_count,
            completed_count=completed_count,
            client_reply_count=client_reply_count,
            manual_pause_count=manual_pause_count,
            total_count=total_count,
 
            unread_replies=get_unread_replies(),
 
            search=""
        )
 
    except Exception as e:
 
        print(
            "Workspace detail error:",
            str(e)
        )
 
        return "Failed to load workspace.", 500
@app.route("/workspace/<int:workspace_id>/remove/<conversation_id>", methods=["POST"])
def remove_from_workspace(workspace_id, conversation_id):
 
    try:
 
        conn = get_connection()
        cursor = conn.cursor()
 
        cursor.execute("""
            DELETE FROM workspace_conversations
            WHERE workspace_id = ?
            AND conversation_id = ?
        """, (
            workspace_id,
            conversation_id
        ))
 
        conn.commit()
        conn.close()
 
        return jsonify({
            "success": True
        })
 
    except Exception as e:
 
        print(
            "Remove from workspace error:",
            str(e)
        )
 
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
 
@app.route(
    "/workspace/<int:workspace_id>/move/<conversation_id>",
    methods=["POST"]
)
def move_workspace_conversation(workspace_id, conversation_id):
 
    try:
 
        data = request.get_json()
 
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body is required."
            }), 400
 
        new_workspace_id = data.get("workspace_id")
 
        if not new_workspace_id:
            return jsonify({
                "success": False,
                "error": "Target workspace is required."
            }), 400
 
        new_workspace_id = int(new_workspace_id)
 
        # Prevent moving to the same workspace
        if new_workspace_id == workspace_id:
            return jsonify({
                "success": False,
                "error": "Conversation is already in this workspace."
            }), 400
 
        # Make sure the conversation actually belongs
        # to the current workspace
        conn = get_connection()
        cursor = conn.cursor()
 
        cursor.execute("""
            SELECT id
            FROM workspace_conversations
            WHERE workspace_id = ?
            AND conversation_id = ?
        """, (
            workspace_id,
            conversation_id
        ))
 
        existing = cursor.fetchone()
 
        conn.close()
 
        if not existing:
            return jsonify({
                "success": False,
                "error": "Conversation is not in this workspace."
            }), 404
 
        # Use the existing DB function
        moved = move_conversation_to_workspace(
            conversation_id,
            new_workspace_id
        )
 
        if not moved:
            return jsonify({
                "success": False,
                "error": "Conversation could not be moved."
            }), 404
 
        return jsonify({
            "success": True
        })
 
    except Exception as e:
 
        print(
            "Move workspace conversation error:",
            str(e)
        )
 
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
@app.route("/workspaces/assign", methods=["POST"])
def assign_conversations_to_workspace():
 
    try:
        data = request.get_json()
 
        if not data:
            return jsonify({
                "error": "Request body is required."
            }), 400
 
        workspace_id = data.get("workspace_id")
        conversation_ids = data.get("conversation_ids", [])

        print("WORKSPACE ASSIGN REQUEST")
        print("Workspace:", workspace_id)
        print("Conversations:", conversation_ids)
 
 
        if not workspace_id:
            return jsonify({
                "error": "Workspace ID is required."
            }), 400
 
        if not conversation_ids:
            return jsonify({
                "error": "At least one conversation is required."
            }), 400
 
        assigned = []
        skipped = []
 
        for conversation_id in conversation_ids:
 
            try:
 
                add_conversation_to_workspace(
                    int(workspace_id),
                    conversation_id
                )
 
                assigned.append(conversation_id)
 
            except Exception as e:
 
                print(
                    f"Could not assign {conversation_id}: {e}"
                )
 
                skipped.append({
                    "conversation_id": conversation_id,
                    "reason": str(e)
                })
 
        return jsonify({
            "success": True,
            "assigned": assigned,
            "skipped": skipped
        }), 200
 
    except Exception as e:
 
        print(
            "Workspace assignment error:",
            str(e)
        )
 
        return jsonify({
            "error": str(e)
        }), 500
 
@app.route("/api/workspaces", methods=["GET"])
def api_get_workspaces():
 
    try:
 
        workspaces = get_workspaces()
 
        return jsonify(workspaces)
 
    except Exception as e:
 
        print("Error loading workspaces:", str(e))
 
        return jsonify({
            "error": str(e)
        }), 500
@app.route("/workspaces/create", methods=["POST"])
def create_workspace_route():
    try:
        data = request.get_json()
 
        if not data:
            return jsonify({
                "error": "Request body is required."
            }), 400
 
        workspace_name = str(
            data.get("workspace_name", "")
        ).strip()
 
        description = str(
            data.get("description", "")
        ).strip()
 
        if not workspace_name:
            return jsonify({
                "error": "Workspace name is required."
            }), 400
 
        workspace_id = create_workspace(
            workspace_name,
            description
        )
 
        return jsonify({
            "success": True,
            "workspace": {
                "id": workspace_id,
                "workspace_name": workspace_name,
                "description": description
            }
        }), 201
 
    except sqlite3.IntegrityError:
        return jsonify({
            "error": "A workspace with this name already exists."
        }), 409
 
    except Exception as e:
        print("Create workspace error:", str(e))
 
        return jsonify({
            "error": str(e)
        }), 500
@app.route("/workspaces/<int:workspace_id>/add-conversation", methods=["POST"])
def add_conversation_to_workspace_route(workspace_id):
 
    try:
        data = request.get_json()
 
        if not data:
            return jsonify({
                "error": "Request body is required."
            }), 400
 
        conversation_id = str(
            data.get("conversation_id", "")
        ).strip()
 
        if not conversation_id:
            return jsonify({
                "error": "Conversation ID is required."
            }), 400
 
        add_conversation_to_workspace(
            workspace_id,
            conversation_id
        )
 
        return jsonify({
            "success": True,
            "message": "Conversation added to workspace."
        })
 
    except Exception as e:
 
        print("Add conversation to workspace error:", str(e))
 
        return jsonify({
            "error": str(e)
        }), 400
 
 
@app.route("/pause/<int:id>")
def pause(id):
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        UPDATE followups
        SET status='MANUAL_PAUSED', updated_at=?
        WHERE id=?
    """, (datetime.utcnow().isoformat(), id))
 
    conn.commit()
    conn.close()

    log_activity(id,"Paused by User")
 
    flash(f"Followup ID {id} paused successfully.")
    return redirect(url_for("dashboard"))
 
 
@app.route("/resume/<int:id>")
def resume(id):
    conn = get_connection()
    cursor = conn.cursor()
 
    now = datetime.utcnow().isoformat()
 
    cursor.execute("""
        UPDATE followups 
        SET status='ACTIVE',
            next_followup_at=?,
            updated_at=? 
        WHERE id=?
    """, (now, now, id))
 
    conn.commit()
    conn.close()

    log_activity(id,"Resumed by User")
 
    flash(f"Followup ID {id} resumed successfully.")
    return redirect(url_for("dashboard"))
 
 
@app.route("/ignore_reply", methods=["POST"])
def ignore_reply_api():
    #data = request.json()
    conversation_id = request.form.get("conversation_id")
 
    if not conversation_id:
        return {"error": "Missing conversation_id"}, 400
 
    ignore_reply(conversation_id)
 
    return redirect("/")

@app.route("/reset/<int:id>")
def reset(id):
    conn = get_connection()
    cursor = conn.cursor()
 
    cursor.execute("""
        UPDATE followups
        SET attempt_count=0,
            status='ACTIVE',
            updated_at=?
        WHERE id=?
    """, (datetime.utcnow().isoformat(), id))
 
    conn.commit()
    conn.close()
 
    flash(f"Followup ID {id} reset successfully.")
    return redirect(url_for("dashboard"))
 
 
# ===============================
# Save Category
# ===============================
# This function is for dashboard form
@app.route("/save_category", methods=["POST"])
def save_category():
    '''print("Form Recieved")
    print(request.form)
    print("--------")
    return "Check terminal"'''
 
    category_name = request.form.get("category_name", "").strip()
    followup_mode = request.form.get("followup_mode", "manual")
    followup_text = request.form.get("followup_text", "").strip()
    max_attempts = request.form.get("max_attempts")
    interval_minutes = request.form.get("interval_minutes")
    selected_draft_ids=request.form.get("selected_draft_ids","").strip()
 
    # Basic required validation
    if not category_name or not max_attempts or not interval_minutes:
        flash("Category name, attempts and interval are required.")
        return redirect(url_for("dashboard"))
 
    # Manual mode requires followup text
    if followup_mode == "manual" and not followup_text:
        flash("Followup text is required in Manual mode.")
        return redirect(url_for("dashboard"))
 
    try:
        save_settings(
            category_name,
            followup_text if followup_mode == "manual" else "",  # store empty if template mode
            int(max_attempts),
            int(interval_minutes),
            followup_mode,
            selected_draft_ids
        )
        token=get_access_token()
        headers={
            "Authorization":f"Bearer {token}",
            "Content-Type":"application/json"
        }

        ensure_outlook_category_exists(category_name,headers)
 
        flash(f"Category '{category_name}' saved successfully and created in Outlook.")
 
    except Exception as e:
        flash(f"Error saving category: {e}")
 
    return redirect(url_for("dashboard"))
 
# ===============================
# Run App
# ===============================

if __name__ == "__main__":
    init_db()
    app.run(debug=True)
     
 